from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse
import datetime
from django.db.models import Sum
from django.contrib.auth.models import User
from collections import defaultdict
import csv
from django.db.models.functions import TruncMonth
from .models import Project, Indicator, MonthlyEntry
from .forms import ProjectForm, IndicatorForm
from django.shortcuts import render
from django.contrib.auth.models import User
from django.db.models import Count
from core.models import Report
from core.templatetags.custom_tags import get_month_name
import calendar
import json
from .models import Indicator
from decimal import Decimal
from django.db.models import F, ExpressionWrapper, FloatField
from django.utils.timezone import now
from datetime import timedelta
from django.db.models import Max
from django.template.loader import render_to_string
from datetime import date
from decimal import Decimal, DivisionUndefined, InvalidOperation
from django.shortcuts import render, get_object_or_404, redirect
from django.forms import modelformset_factory
from django.utils import timezone
from .models import Indicator, IndicatorTarget
from .forms import IndicatorTargetForm #from django.shortcuts import render
from decimal import Decimal, InvalidOperation
from django.db import transaction, IntegrityError
import datetime
from collections import defaultdict
from .forms import IndicatorForm
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image  # for logo
import os
from django.conf import settings
from django.db import models

def safe_int(value, default=0):
    """
    Safely converts a value to integer.
    Returns `default` if value is None, empty string, or invalid.
    """
    try:
        if value in [None, '']:
            return default
        return int(value)
    except (ValueError, TypeError):
        return default


# -----------------------------
# Dashboard
# -----------------------------

from django.db.models import Sum, Prefetch, Q
from .utils import safe_int
import re

@login_required
def dashboard(request):
    """Dashboard view showing KPIs, comparative analysis, trend analysis, rotating indicators,
    gamification stats, system-wide highlights/risks, and district/facility coverage."""

    today_year = datetime.date.today().year
    selected_year = safe_int(request.GET.get("year"), default=today_year)

    # -----------------------------
    # Available years for dropdown
    # -----------------------------
    years_set = set(
        list(MonthlyEntry.objects.values_list("year", flat=True).distinct()) +
        list(IndicatorTarget.objects.values_list("year", flat=True).distinct()) +
        [today_year]
    )
    years = sorted(years_set, reverse=True)

    # -----------------------------
    # Active projects & indicators
    # -----------------------------
    projects = Project.objects.filter(is_active=True)
    total_projects = projects.count()
    all_indicators = Indicator.objects.filter(project__is_active=True, is_active=True).select_related("project")
    total_indicators = all_indicators.count()
    total_entries = MonthlyEntry.objects.filter(year=selected_year, indicator__is_active=True).count()

    # -----------------------------
    # Annual totals & users
    # -----------------------------
    annual_total = safe_int(
        MonthlyEntry.objects.filter(
            year=selected_year,
            indicator__is_active=True,
            indicator__unit__iexact="People Reached"
        ).aggregate(total=Sum("value"))["total"]
    )
    user_count = User.objects.filter(is_active=True).count()

    # -----------------------------
    # System-wide KPI performance
    # -----------------------------
    system_achieved = system_pending = system_behind = 0
    indicator_entries = MonthlyEntry.objects.filter(
        year=selected_year,
        indicator__is_active=True
    ).values('indicator').annotate(total=Sum('value'))
    indicator_totals = {entry['indicator']: entry['total'] for entry in indicator_entries}

    for ind in all_indicators:
        total_actual = safe_int(indicator_totals.get(ind.id, 0))
        target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first()
        target_value = safe_int(target_obj.value) if target_obj else 0

        if target_value == 0:
            system_pending += 1
            continue

        ratio = (total_actual / target_value) * 100
        if ratio >= 90:
            system_achieved += 1
        elif ratio >= 60:
            system_pending += 1
        else:
            system_behind += 1

    system_progress = round((system_achieved / total_indicators) * 100, 2) if total_indicators else 0
    overall_progress = system_progress

    # -----------------------------
    # Overall Trend Calculation (All KPIs)
    # -----------------------------
    monthly_totals = (
        MonthlyEntry.objects.filter(year=selected_year, indicator__is_active=True)
        .values("month")
        .annotate(total=Sum("value"))
        .order_by("month")
    )
    monthly_dict = {entry["month"]: safe_int(entry["total"]) for entry in monthly_totals}
    last_month = max(monthly_dict.keys()) if monthly_dict else 0
    last_value = monthly_dict.get(last_month, 0)
    prev_months = [m for m in monthly_dict.keys() if m < last_month]
    prev_value = monthly_dict[max(prev_months)] if prev_months else 0

    if prev_value == 0 and last_value == 0:
        overall_trend = 0
    elif prev_value == 0 and last_value > 0:
        overall_trend = 100
    else:
        overall_trend = round(((last_value - prev_value) / prev_value) * 100, 2)

    # -----------------------------
    # System-wide Highlights & Risks
    # -----------------------------
    highlights, risks = [], []

    for ind in all_indicators:
        total_actual = safe_int(indicator_totals.get(ind.id, 0))
        target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first()
        target_value = safe_int(target_obj.value) if target_obj else None
        ratio = (total_actual / target_value * 100) if target_value else 0

        if target_value and ratio >= 90:
            highlights.append({
                "name": ind.name,
                "project": ind.project.name if ind.project else "—",
                "value": total_actual,
                "target": target_value,
                "ratio": round(ratio, 1),
            })
        if target_value and ratio < 60:
            risks.append({
                "name": ind.name,
                "project": ind.project.name if ind.project else "—",
                "value": total_actual,
                "target": target_value,
                "ratio": round(ratio, 1),
            })

    if not highlights:
        temp = []
        for ind in all_indicators:
            total_actual = safe_int(indicator_totals.get(ind.id, 0))
            target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first()
            target_value = safe_int(target_obj.value) if target_obj else 0
            ratio = (total_actual / target_value * 100) if target_value else 0
            temp.append({
                "name": ind.name,
                "project": ind.project.name if ind.project else "—",
                "value": total_actual,
                "target": target_value,
                "ratio": round(ratio, 1),
            })
        highlights = sorted(temp, key=lambda x: x['ratio'], reverse=True)[:3]

    if not risks:
        temp = []
        for ind in all_indicators:
            total_actual = safe_int(indicator_totals.get(ind.id, 0))
            target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first()
            target_value = safe_int(target_obj.value) if target_obj else 0
            ratio = (total_actual / target_value * 100) if target_value else 0
            temp.append({
                "name": ind.name,
                "project": ind.project.name if ind.project else "—",
                "value": total_actual,
                "target": target_value,
                "ratio": round(ratio, 1),
            })
        risks = sorted(temp, key=lambda x: x['ratio'])[:3]

    # -----------------------------
    # User projects & gamification
    # -----------------------------
    user_projects = Project.objects.filter(coordinator=request.user, is_active=True)
    if not user_projects.exists():
        user_projects = Project.objects.filter(
            coordinator__first_name=request.user.first_name,
            coordinator__last_name=request.user.last_name,
            is_active=True
        )

    user_projects_count = user_projects.count()
    user_project_indicators = Indicator.objects.filter(project__in=user_projects, is_active=True)
    user_total_indicators = user_project_indicators.count()
    user_kpi_url = f"/projects/{user_projects.first().id}/kpis/" if user_projects.exists() else "#"

    user_achieved = user_pending = user_behind = 0
    user_entries = MonthlyEntry.objects.filter(
        year=selected_year,
        created_by=request.user,
        indicator__in=user_project_indicators
    ).values('indicator').annotate(total=Sum('value'))
    user_totals = {entry['indicator']: entry['total'] for entry in user_entries}

    for ind in user_project_indicators:
        user_total_actual = safe_int(user_totals.get(ind.id, 0))
        target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first()
        target_value = safe_int(target_obj.value) if target_obj else 0

        if target_value == 0:
            user_pending += 1
        else:
            ratio = (user_total_actual / target_value) * 100
            if ratio >= 90:
                user_achieved += 1
            elif ratio >= 60:
                user_pending += 1
            else:
                user_behind += 1

    user_progress = round((user_achieved / user_total_indicators) * 100, 2) if user_total_indicators else 0

    # -----------------------------
    # Comparative & Trend Analysis (People Reached only)
    # -----------------------------
    comparative_labels, comparative_actuals, comparative_targets = [], [], []
    for proj in projects:
        comparative_labels.append(proj.name)
        active_indicators = Indicator.objects.filter(project=proj, is_active=True)
        actual_total = safe_int(
            MonthlyEntry.objects.filter(
                year=selected_year,
                indicator__in=active_indicators,
                indicator__unit__iexact="People Reached"
            ).aggregate(total=Sum("value"))["total"]
        )
        indicator_total = safe_int(
            IndicatorTarget.objects.filter(
                indicator__in=active_indicators,
                year=selected_year,
                indicator__unit__iexact="People Reached"
            ).aggregate(total=Sum("value"))["total"]
        )
        comparative_actuals.append(actual_total)
        comparative_targets.append(indicator_total)

    trend_months, trend_actuals, trend_targets = [], [], []
    total_annual_target = safe_int(
        IndicatorTarget.objects.filter(
            year=selected_year,
            indicator__unit__iexact="People Reached",
            indicator__is_active=True
        ).aggregate(total=Sum("value"))["total"]
    )
    monthly_target = int(total_annual_target / 12) if total_annual_target else 0
    for month_num in range(1, 13):
        trend_months.append(datetime.date(1900, month_num, 1).strftime("%b"))
        monthly_actual = safe_int(
            MonthlyEntry.objects.filter(
                year=selected_year,
                month=month_num,
                indicator__unit__iexact="People Reached",
                indicator__is_active=True
            ).aggregate(total=Sum("value"))["total"]
        )
        trend_actuals.append(monthly_actual)
        trend_targets.append(monthly_target)

    # -----------------------------
    # Rotating Indicators
    # -----------------------------
    indicators_data = []
    for ind in all_indicators:
        total_actual = safe_int(indicator_totals.get(ind.id, 0))
        last_entry = MonthlyEntry.objects.filter(indicator=ind, year=selected_year).order_by("-id").first()
        prev_entry_qs = MonthlyEntry.objects.filter(indicator=ind, year=selected_year).order_by("-id")[1:2]
        previous_value = safe_int(prev_entry_qs[0].value) if prev_entry_qs else 0
        last_value = safe_int(last_entry.value) if last_entry else 0
        target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first()
        target_value = safe_int(target_obj.value) if target_obj else 0
        indicators_data.append({
            "id": ind.id,
            "name": ind.name,
            "project": ind.project.name if ind.project else None,
            "value": total_actual,
            "target": target_value,
            "previous_value": previous_value,
            "last_month_value": last_value,
        })

    # -----------------------------
    # Districts & Facilities Coverage (ACTIVE ONLY)
    # -----------------------------
    districts = District.objects.filter(is_active=True)
    facilities = Facility.objects.filter(is_active=True, projects__is_active=True).distinct()
    total_facilities = facilities.count()
    district_coverage = []

    for district in districts:
        district_facilities_count = facilities.filter(district=district).count()
        percentage = round((district_facilities_count / total_facilities) * 100, 2) if total_facilities else 0
        district_coverage.append({
            "district": district,
            "facility_count": district_facilities_count,
            "percentage": percentage,
            "display": f"{district.name} ({percentage}%)"
        })

    total_covered_facilities = sum(d['facility_count'] for d in district_coverage)
    total_coverage_percent = round((total_covered_facilities / total_facilities) * 100, 2) if total_facilities else 0

    # -----------------------------
    # ✅ Gender KPI Calculations (Word-boundary matching)
    # -----------------------------
    male_terms = ["male", "males", "man", "men", "boy", "boys"]
    female_terms = ["female", "females", "woman", "women", "girl", "girls", "agyw", "agyws"]

    male_regex = re.compile(r'\b(' + '|'.join(male_terms) + r')\b', re.IGNORECASE)
    female_regex = re.compile(r'\b(' + '|'.join(female_terms) + r')\b', re.IGNORECASE)

    gender_counts = {"male": 0, "female": 0, "neutral": 0}

    for ind in all_indicators:
        name = ind.name.lower().strip()
        if female_regex.search(name):
            gender_counts["female"] += 1
        elif male_regex.search(name):
            gender_counts["male"] += 1
        else:
            gender_counts["neutral"] += 1

    total_gender = sum(gender_counts.values()) or 1
    overall_gender = {
        "female_count": gender_counts["female"],
        "male_count": gender_counts["male"],
        "neutral_count": gender_counts["neutral"],
        "total_gender_kpis": total_gender,
        "female_percent": round(gender_counts["female"] / total_gender * 100, 1),
        "male_percent": round(gender_counts["male"] / total_gender * 100, 1),
        "neutral_percent": round(gender_counts["neutral"] / total_gender * 100, 1),
    }

    # -----------------------------
    # KPI Summary Cards
    # -----------------------------
    stats = [
        {"label": "Projects", "value": total_projects, "icon": "fas fa-folder-open", "bg_gradient": "linear-gradient(135deg, #6a11cb, #2575fc)", "url": "/projects/"},
        {"label": "Indicators", "value": total_indicators, "icon": "fas fa-bullseye", "bg_gradient": "linear-gradient(135deg, #11998e, #38ef7d)", "url": "/projects/"},
        {"label": f"People Reached ({selected_year})", "value": annual_total, "icon": "fas fa-calendar-alt", "bg_gradient": "linear-gradient(135deg, #f7971e, #ffd200)", "url": "#"},
        {"label": "Active Users", "value": user_count, "icon": "fas fa-users", "bg_gradient": "linear-gradient(135deg, #ff416c, #ff4b2b)", "url": "/profile/"},
    ]

    # -----------------------------
    # Leaderboard
    # -----------------------------
    user_points = defaultdict(int)
    user_badges_dict = defaultdict(list)
    for proj in Project.objects.filter(is_active=True):
        coordinator = proj.coordinator
        if not coordinator:
            continue
        active_indicators = Indicator.objects.filter(project=proj, is_active=True)
        total_value = MonthlyEntry.objects.filter(year=selected_year, indicator__in=active_indicators).aggregate(total=Sum('value'))['total'] or 0
        user_points[coordinator.id] += total_value
        if total_value >= 1000:
            user_badges_dict[coordinator.id].append("Gold")
        elif total_value >= 500:
            user_badges_dict[coordinator.id].append("Silver")
        elif total_value >= 100:
            user_badges_dict[coordinator.id].append("Bronze")

    top_users = []
    for uid, points in user_points.items():
        user_obj = User.objects.filter(id=uid).first()
        if user_obj:
            top_users.append({
                "username": f"{user_obj.first_name} {user_obj.last_name}".strip() or user_obj.username,
                "total_points": points,
                "badges": user_badges_dict.get(uid, []),
            })
    top_users = sorted(top_users, key=lambda x: x['total_points'], reverse=True)[:3]

    # -----------------------------
    # User Badges
    # -----------------------------
    user_badges = []
    if user_total_indicators:
        if user_progress >= 90:
            user_badges.append("Gold")
        elif user_progress >= 60:
            user_badges.append("Silver")
        elif user_progress > 0:
            user_badges.append("Bronze")
        else:
            user_badges.append("No Badge")

    # -----------------------------
    # Context
    # -----------------------------
    context = {
        "projects": projects,
        "total_projects": total_projects,
        "total_indicators": total_indicators,
        "total_entries": total_entries,
        "current_year": today_year,
        "selected_year": selected_year,
        "years": years,
        "annual_total": annual_total,
        "user_count": user_count,
        "comparative_labels": comparative_labels,
        "comparative_actuals": comparative_actuals,
        "comparative_targets": comparative_targets,
        "trend_months": trend_months,
        "trend_actuals": trend_actuals,
        "trend_targets": trend_targets,
        "indicators": indicators_data,
        "districts": districts,
        "facilities": facilities,
        "district_coverage": district_coverage,
        "total_coverage_percent": total_coverage_percent,
        "stats": stats,
        "system_achieved": system_achieved,
        "system_pending": system_pending,
        "system_behind": system_behind,
        "system_progress": system_progress,
        "overall_progress": overall_progress,
        "overall_trend": overall_trend,
        "user_achieved": user_achieved,
        "user_pending": user_pending,
        "user_behind": user_behind,
        "user_progress": user_progress,
        "highlights": highlights,
        "risks": risks,
        "user_projects_count": user_projects_count,
        "user_total_indicators": user_total_indicators,
        "user_kpi_url": user_kpi_url,
        "top_users": top_users,
        "user_badges": user_badges,
        "gender_counts": gender_counts,
        "overall_gender": overall_gender,
    }

    return render(request, "core/dashboard.html", context)




@login_required
def project_detail(request, pk):
    """
    Show project details with monthly KPI grid.
    Only includes active projects and active indicators.
    """
    # Ensure project is active
    project = get_object_or_404(Project.objects.filter(is_active=True), pk=pk)

    # Only active indicators for this project
    indicators = Indicator.objects.filter(project=project, is_active=True)

    months = list(range(1, 13))
    year = int(request.GET.get("year", datetime.datetime.now().year))

    # --- Build KPI grid ---
    grid = {}
    for ind in indicators:
        grid[ind.id] = {}
        for month in months:
            entry = MonthlyEntry.objects.filter(
                indicator=ind, year=year, month=month
            ).first()
            grid[ind.id][month] = entry.value if entry else None

    return render(request, "core/project_detail.html", {
        "project": project,
        "indicators": indicators,
        "months": months,
        "year": year,
        "grid": grid,
    })


# -----------------------------
# KPI Data Entry View (Strict Numeric Validation)
# -----------------------------

@login_required
def project_kpis(request, pk):
    """
    KPI Data entry for a project.
    Only active projects and active indicators are included.
    Handles monthly entries and optional annual target updates.
    Strictly rejects non-numeric values (letters, commas, dots, symbols).
    """
    # Ensure project is active
    project = get_object_or_404(Project.objects.filter(is_active=True), pk=pk)

    # Only active indicators
    indicators = Indicator.objects.filter(project=project, is_active=True)

    months = list(range(1, 13))
    year = int(request.GET.get("year", request.POST.get("year", datetime.datetime.now().year)))

    errors = defaultdict(dict)  # {indicator_id: {month: error_message}}

    # -------------------------
    # Handle POST: bulk monthly values + optional target updates
    # -------------------------
    if request.method == "POST":
        with transaction.atomic():
            for ind in indicators:
                # --- Update monthly entries ---
                for month in months:
                    field_name = f"val_{ind.id}_{month}"
                    raw_val = request.POST.get(field_name, "").strip()
                    if raw_val == "":
                        continue
                    if not raw_val.isdigit():
                        errors[ind.id][month] = "Only whole numbers allowed"
                        continue
                    val = int(raw_val)
                    MonthlyEntry.objects.update_or_create(
                        indicator=ind,
                        year=year,
                        month=month,
                        defaults={"value": val, "created_by": request.user},
                    )

                # --- Update annual target if provided ---
                target_field = f"target_{ind.id}"
                raw_target = request.POST.get(target_field, "").strip()
                if raw_target != "":
                    if not raw_target.isdigit():
                        errors[ind.id]["target"] = "Only whole numbers allowed"
                        continue
                    target_val = int(raw_target)
                    IndicatorTarget.objects.update_or_create(
                        indicator=ind,
                        year=year,
                        defaults={"value": target_val},
                    )

                # --- Update indicator progress ---
                total_val = MonthlyEntry.objects.filter(indicator=ind, year=year).aggregate(
                    total=Sum("value")
                )["total"] or 0
                ind.current_value = total_val
                ind.update_progress()

        # -------------------------
        # Handle messages and redirect
        # -------------------------
        if any(errors[ind.id] for ind in indicators):
            messages.warning(request, "Some cells contained invalid values. Only whole numbers are allowed. Please correct them.")
        else:
            messages.success(request, "KPI data and targets saved successfully!")
            return redirect("project_kpis", pk=project.id)

    # -------------------------
    # Prepare grid + totals
    # -------------------------
    grid = defaultdict(dict)
    totals = {}              # Row totals per indicator
    monthly_totals = {}      # Column totals across indicators
    grand_total = 0          # Bottom-right grand total

    # Fetch targets for all indicators for this year
    targets = IndicatorTarget.objects.filter(indicator__in=indicators, year=year)
    target_dict = {t.indicator_id: int(t.value) for t in targets}

    # --- Auto-create target for new indicators ---
    for ind in indicators:
        if ind.id not in target_dict:
            target_obj, created = IndicatorTarget.objects.get_or_create(
                indicator=ind,
                year=year,
                defaults={"value": 0},  # integer default
            )
            target_dict[ind.id] = int(target_obj.value)

    # Build grid and row totals
    for ind in indicators:
        row_total = 0
        for month in months:
            entry = MonthlyEntry.objects.filter(indicator=ind, year=year, month=month).first()
            val = int(entry.value) if entry else 0
            grid[ind.id][month] = val if val != 0 else None
            row_total += val
        totals[ind.id] = row_total
        grand_total += row_total

    # Build column totals
    for month in months:
        monthly_totals[month] = sum(grid[ind.id].get(month, 0) or 0 for ind in indicators)

    context = {
        "project": project,
        "indicators": indicators,
        "months": months,
        "year": year,
        "grid": grid,
        "totals": totals,
        "monthly_totals": monthly_totals,
        "grand_total": grand_total,
        "target_dict": target_dict,
        "errors": errors,  # Pass errors to template
    }

    return render(request, "core/project_kpis.html", context)




# -------------------------
# ADD PROJECT (Multiple Donors + Coordinator + Districts)
# -------------------------
from django.contrib.auth import get_user_model

User = get_user_model()
@login_required
def project_add(request):
    all_donors = Donor.objects.all().order_by('name')
    all_facilities = Facility.objects.all().order_by('name')
    all_districts = District.objects.all().order_by('name')
    all_users = User.objects.filter(is_active=True).order_by('username')

    form = ProjectForm(request.POST or None)

    # Precompute selected values for template
    selected_donors = request.POST.getlist('donors') if request.method == "POST" else []
    selected_facilities = request.POST.getlist('facilities') if request.method == "POST" else []
    selected_districts = request.POST.getlist('districts') if request.method == "POST" else []
    selected_coordinator = request.POST.get('coordinator') if request.method == "POST" else None
    selected_main_donor = request.POST.get('main_donor') if request.method == "POST" else None

    if request.method == "POST" and form.is_valid():
        project = form.save(commit=False)

        # Assign coordinator
        project.coordinator = User.objects.filter(id=selected_coordinator).first() if selected_coordinator else None

        project.save()

        # Assign many-to-many fields
        project.donors.set(selected_donors)
        project.facilities.set(selected_facilities)
        project.districts.set(selected_districts)

        messages.success(request, "Project added successfully!")
        return redirect("projects")

    context = {
        "form": form,
        "title": "Add Project",
        "all_donors": all_donors,
        "all_facilities": all_facilities,
        "all_districts": all_districts,
        "all_users": all_users,
        "project": None,
        "selected_donors": selected_donors,
        "selected_facilities": selected_facilities,
        "selected_districts": selected_districts,
        "selected_coordinator": selected_coordinator,
        "selected_main_donor": selected_main_donor,
    }
    return render(request, "core/project_form.html", context)


@login_required
def project_edit(request, pk):
    project = get_object_or_404(Project, pk=pk)
    all_donors = Donor.objects.all().order_by('name')
    all_facilities = Facility.objects.all().order_by('name')
    all_districts = District.objects.all().order_by('name')
    all_users = User.objects.filter(is_active=True).order_by('username')

    form = ProjectForm(request.POST or None, instance=project)

    # Precompute selected values for template
    selected_donors = request.POST.getlist('donors') if request.method == "POST" else [str(d.id) for d in project.donors.all()]
    selected_facilities = request.POST.getlist('facilities') if request.method == "POST" else [str(f.id) for f in project.facilities.all()]
    selected_districts = request.POST.getlist('districts') if request.method == "POST" else [str(d.id) for d in project.districts.all()]
    selected_coordinator = request.POST.get('coordinator') if request.method == "POST" else (str(project.coordinator.id) if project.coordinator else None)
    selected_main_donor = request.POST.get('main_donor') if request.method == "POST" else (str(project.main_donor.id) if project.main_donor else None)

    if request.method == "POST" and form.is_valid():
        project = form.save(commit=False)

        # Assign coordinator
        project.coordinator = User.objects.filter(id=selected_coordinator).first() if selected_coordinator else None

        project.save()

        # Assign many-to-many fields
        project.donors.set(selected_donors)
        project.facilities.set(selected_facilities)
        project.districts.set(selected_districts)

        messages.success(request, "Project updated successfully!")
        return redirect("projects")

    context = {
        "form": form,
        "title": "Edit Project",
        "all_donors": all_donors,
        "all_facilities": all_facilities,
        "all_districts": all_districts,
        "all_users": all_users,
        "project": project,
        "selected_donors": selected_donors,
        "selected_facilities": selected_facilities,
        "selected_districts": selected_districts,
        "selected_coordinator": selected_coordinator,
        "selected_main_donor": selected_main_donor,
    }
    return render(request, "core/project_form.html", context)





@login_required
def project_delete(request, pk):
    project = get_object_or_404(Project, pk=pk)
    if request.method == "POST":
        project.delete()
        messages.success(request, "Project deleted successfully!")
        return redirect("projects")
    return render(request, "core/project_confirm_delete.html", {"project": project})




# -----------------------------
# Add KPI view
# -----------------------------
@login_required
def indicator_add(request, project_pk):
    project = get_object_or_404(Project, pk=project_pk)
    current_year = timezone.now().year
    years = [current_year - 2, current_year - 1, current_year, current_year + 1, current_year + 2]

    if request.method == "POST":
        form = IndicatorForm(request.POST)
        if form.is_valid():
            # Save Indicator instance first
            indicator = form.save(commit=False)
            indicator.project = project
            indicator.save()

            # Handle KPI target for selected year
            target_val_raw = form.cleaned_data.get("target")
            target_year_raw = request.POST.get("year", current_year)

            try:
                target_year = int(target_year_raw)
            except (ValueError, TypeError):
                target_year = current_year

            # Validate numeric target before saving
            if target_val_raw is not None and str(target_val_raw).strip() != "":
                try:
                    target_val = safe_int(target_val_raw)
                except ValueError:
                    messages.warning(
                        request,
                        "Invalid target value. Only whole numbers are allowed."
                    )
                    return render(
                        request,
                        "core/indicators/indicator_form.html",
                        {
                            "form": form,
                            "project": project,
                            "title": "Add KPI",
                            "now": timezone.now(),
                            "years": years,
                            "selected_year": target_year,
                        },
                    )

                IndicatorTarget.objects.update_or_create(
                    indicator=indicator,
                    year=target_year,
                    defaults={"value": target_val},
                )
            else:
                # If target is empty, default to 0
                IndicatorTarget.objects.update_or_create(
                    indicator=indicator,
                    year=target_year,
                    defaults={"value": 0},
                )

            messages.success(request, "KPI added successfully!")
            return redirect("project_detail", pk=project.pk)
        else:
            messages.warning(request, "Please correct the errors in the form.")
    else:
        form = IndicatorForm()

    return render(
        request,
        "core/indicators/indicator_form.html",
        {
            "form": form,
            "project": project,
            "title": "Add KPI",
            "now": timezone.now(),
            "years": years,
            "selected_year": current_year,
        },
    )




# -----------------------------
# Edit KPI view
# -----------------------------
@login_required
def indicator_edit(request, project_pk, pk):
    project = get_object_or_404(Project, pk=project_pk)
    indicator = get_object_or_404(Indicator, pk=pk, project=project)
    current_year = timezone.now().year
    years = [current_year - 2, current_year - 1, current_year, current_year + 1, current_year + 2]

    # Pre-fill current-year target if it exists
    try:
        current_target = IndicatorTarget.objects.get(indicator=indicator, year=current_year)
        initial_data = {"target": current_target.value}
        selected_year = current_year
    except IndicatorTarget.DoesNotExist:
        initial_data = {}
        selected_year = current_year

    if request.method == "POST":
        form = IndicatorForm(request.POST, instance=indicator, initial=initial_data)
        if form.is_valid():
            indicator = form.save(commit=False)
            indicator.project = project
            indicator.save()

            # Handle KPI target for selected year
            target_val_raw = form.cleaned_data.get("target")
            target_year_raw = request.POST.get("year", current_year)

            try:
                target_year = int(target_year_raw)
            except (ValueError, TypeError):
                target_year = current_year

            # Validate numeric target before saving
            if target_val_raw is not None and str(target_val_raw).strip() != "":
                try:
                    target_val = safe_int(target_val_raw)
                    IndicatorTarget.objects.update_or_create(
                        indicator=indicator,
                        year=target_year,
                        defaults={"value": target_val},
                    )
                    messages.success(request, "KPI updated successfully!")
                    return redirect("project_detail", pk=project.pk)
                except ValueError:
                    messages.warning(
                        request,
                        "Invalid target value. Only whole numbers are allowed."
                    )
            else:
                # If target is empty, default to 0
                IndicatorTarget.objects.update_or_create(
                    indicator=indicator,
                    year=target_year,
                    defaults={"value": 0},
                )
                messages.success(request, "KPI updated successfully!")
                return redirect("project_detail", pk=project.pk)
        else:
            messages.warning(request, "Please correct the errors in the form.")
    else:
        form = IndicatorForm(instance=indicator, initial=initial_data)

    return render(
        request,
        "core/indicators/indicator_form.html",
        {
            "form": form,
            "project": project,
            "indicator": indicator,
            "title": "Edit KPI",
            "now": timezone.now(),
            "years": years,
            "selected_year": selected_year,
        },
    )


@login_required
def indicator_delete(request, project_pk, pk):
    project = get_object_or_404(Project, pk=project_pk)
    indicator = get_object_or_404(Indicator, pk=pk, project=project)
    if request.method == "POST":
        indicator.delete()
        messages.success(request, "KPI deleted successfully!")
        return redirect("project_detail", pk=project.pk)
    return render(request, "core/indicators/indicator_confirm_delete.html",
                  {"project": project, "indicator": indicator})


# views.py
from .models import Project

# -------------------------
# USER PROFILE
# -------------------------
@login_required
def profile(request):
    user = request.user

    # -------------------------
    # Fetch projects coordinated by this user
    # -------------------------
    # Option 1 (Preferred): Match by user FK directly
    projects_coordinated = (
        Project.objects.filter(coordinator=user)
        .select_related('main_donor')      # reduce DB queries for main_donor
        .prefetch_related('facilities')    # reduce DB queries for facilities
        .order_by('name')                  # sort alphabetically
    )

    # Option 2: Match by full name (first + last name)
    if not projects_coordinated.exists():
        projects_coordinated = (
            Project.objects.filter(
                coordinator__first_name=user.first_name,
                coordinator__last_name=user.last_name
            )
            .select_related('main_donor')
            .prefetch_related('facilities')
            .order_by('name')
        )

    # Count to simplify template conditional checks
    projects_count = projects_coordinated.count()

    context = {
        'user': user,
        'projects_coordinated': projects_coordinated,
        'projects_count': projects_count,
    }

    return render(request, "core/profile.html", context)




# -------------------------
# LIST PROJECTS WITH FACILITIES, DISTRICTS & MULTIPLE DONORS
# -------------------------

@login_required
def projects(request):
    """
    List all projects with their facilities, districts, donor info, and coordinator.
    Includes status messages: Active, Ending Soon, Completed, Overdue.
    Handles AJAX or standard POST for facility or district updates.
    Supports multi-donor display (main + other donors).
    """
    projects_qs = Project.objects.all().order_by('name')
    all_facilities = Facility.objects.all().order_by('name')
    all_donors = Donor.objects.all().order_by('name')
    all_districts = District.objects.all().order_by('name')

    # -----------------------------
    # Handle updates (AJAX or form)
    # -----------------------------
    if request.method == "POST":
        # JSON/AJAX request
        if request.content_type == "application/json":
            try:
                data = json.loads(request.body)
                project_id = data.get("project_id")
                project = get_object_or_404(Project, pk=project_id)

                # Update facilities if provided
                facilities_ids = data.get("facilities")
                if facilities_ids is not None:
                    project.facilities.set(facilities_ids)

                    # Auto-link districts from selected facilities if not already set
                    facility_districts = District.objects.filter(facilities__in=project.facilities.all()).distinct()
                    project.districts.set(facility_districts)

                # Update districts if provided
                districts_ids = data.get("districts")
                if districts_ids is not None:
                    project.districts.set(districts_ids)

                project.save()
                return JsonResponse({"status": "success"})
            except (json.JSONDecodeError, KeyError):
                return JsonResponse({"status": "error", "message": "Invalid JSON"}, status=400)

        # Standard POST from forms
        if "facilities_submit" in request.POST or "districts_submit" in request.POST:
            project_id = request.POST.get("project_id")
            project = get_object_or_404(Project, pk=project_id)

            if "facilities_submit" in request.POST:
                selected_facilities = request.POST.getlist("facilities")
                project.facilities.set(selected_facilities)

                # Auto-link districts from selected facilities if not already set
                facility_districts = District.objects.filter(facilities__in=project.facilities.all()).distinct()
                project.districts.set(facility_districts)

                messages.success(request, f"Facilities and districts updated for project '{project.name}'.")

            if "districts_submit" in request.POST:
                selected_districts = request.POST.getlist("districts")
                project.districts.set(selected_districts)
                messages.success(request, f"Districts updated for project '{project.name}'.")

            project.save()
            return redirect('projects')

    # -----------------------------
    # Prepare projects for display
    # -----------------------------
    today = date.today()
    projects_with_status = []

    for proj in projects_qs:
        # Determine project status
        status_text = "Active"
        if proj.end_date:
            if proj.end_date < today:
                status_text = "Completed"
            elif (proj.end_date - today).days <= 30:
                status_text = "Ending Soon"

        if proj.end_date and proj.end_date < today and not proj.is_active:
            status_text = "Overdue"

        # Coordinator info
        coordinator_info = None
        if hasattr(proj, 'coordinator') and proj.coordinator:
            coordinator_info = {
                "first_name": proj.coordinator.first_name,
                "last_name": proj.coordinator.last_name,
                "is_active": proj.coordinator.is_active,
            }

        # Districts info
        districts_qs = proj.districts.all().order_by("name")
        district_ids = list(districts_qs.values_list("id", flat=True))  # for template checkbox preselection

        projects_with_status.append({
            "project": proj,
            "status_text": status_text,
            "donors": proj.donors.all(),
            "coordinator": coordinator_info,
            "facilities": proj.facilities.all(),
            "districts": districts_qs,
            "district_ids": district_ids,  # for template
        })

    context = {
        "projects_with_status": projects_with_status,
        "all_facilities": all_facilities,
        "all_districts": all_districts,
        "all_donors": all_donors,
    }

    return render(request, "core/projects.html", context)



# -----------------------------
# Bulk Save Entries View
# -----------------------------

@login_required
def bulk_save_entries(request, pk):
    """
    Allows bulk saving of monthly KPI entries for a project.
    Only active projects and active indicators are included.
    Rejects any non-numeric input (commas, dots, letters, symbols).
    """
    # Ensure project is active
    project = get_object_or_404(Project.objects.filter(is_active=True), pk=pk)

    # Only active indicators for the project
    indicators = Indicator.objects.filter(project=project, is_active=True)

    now = timezone.now()
    current_year = now.year
    current_month = now.month

    if request.method == "POST":
        year = int(request.POST.get("year", current_year))
        month = int(request.POST.get("month", current_month))
        invalid_entries = []

        with transaction.atomic():
            for indicator in indicators:
                raw_value = request.POST.get(f"indicator_{indicator.id}")
                if raw_value not in [None, ""]:
                    raw_value = raw_value.strip()
                    # Reject if not a whole number
                    if not raw_value.isdigit():
                        invalid_entries.append(indicator.name)
                        continue

                    numeric_value = int(raw_value)
                    MonthlyEntry.objects.update_or_create(
                        indicator=indicator,
                        year=year,
                        month=month,
                        defaults={"value": numeric_value, "created_by": request.user},
                    )
                    # Update indicator progress after entry
                    indicator.update_progress()

        if invalid_entries:
            messages.warning(
                request,
                f"Some entries were not saved because they were not valid whole numbers: {', '.join(invalid_entries)}"
            )
        else:
            messages.success(request, f"Entries for {month}/{year} saved successfully!")

        return redirect("project_detail", pk=project.id)

    return render(request, "core/bulk_save_entries.html", {
        "project": project,
        "indicators": indicators,
        "year": current_year,
        "month": current_month,
    })



# -------------------------
# REPORTS
# -------------------------
@login_required
def reports(request):
    current_year = datetime.datetime.now().year
    year = int(request.GET.get("year", current_year))

    # ✅ Collect distinct years available in MonthlyEntry + current year
    years_qs = MonthlyEntry.objects.values_list("year", flat=True).distinct()
    years = sorted(set(list(years_qs) + [current_year]), reverse=True)

    projects = Project.objects.all()
    indicators = Indicator.objects.all()
    units = indicators.values_list("unit", flat=True).distinct()

    selected_unit = request.GET.get("unit") or None
    selected_indicator_ids = request.GET.getlist("indicators")

    base_qs = MonthlyEntry.objects.filter(year=year)
    if selected_unit:
        base_qs = base_qs.filter(indicator__unit=selected_unit)
    if selected_indicator_ids:
        base_qs = base_qs.filter(indicator_id__in=selected_indicator_ids)

    # -------------------------
    # Indicator Totals
    # -------------------------
    indicator_totals = (
        base_qs.values("indicator__id", "indicator__name", "indicator__unit")
        .annotate(total=Sum("value"))
        .order_by("indicator__name")
    )
    selected_indicators_data, overall_total = [], 0
    for row in indicator_totals:
        val = row["total"] or 0
        selected_indicators_data.append({
            "id": row["indicator__id"],
            "name": row["indicator__name"],
            "unit": row["indicator__unit"],
            "value": val
        })
        overall_total += val

    # -------------------------
    # Monthly Breakdown
    # -------------------------
    monthly_labels = list(range(1, 13))
    monthly_data = []
    for ind in indicators:
        ind_series = [
            base_qs.filter(indicator=ind, month=m).aggregate(total=Sum("value"))["total"] or 0
            for m in monthly_labels
        ]
        if any(ind_series):
            monthly_data.append({
                "indicator": ind.name,
                "unit": ind.unit,
                "values": ind_series,
            })

    # -------------------------
    # Quarterly Breakdown
    # -------------------------
    quarterly_labels = ["Q1", "Q2", "Q3", "Q4"]
    quarterly_data = []
    for ind in indicators:
        ind_monthly = [
            base_qs.filter(indicator=ind, month=m).aggregate(total=Sum("value"))["total"] or 0
            for m in monthly_labels
        ]
        if any(ind_monthly):
            quarterly_data.append({
                "indicator": ind.name,
                "unit": ind.unit,
                "values": [
                    sum(ind_monthly[0:3]),
                    sum(ind_monthly[3:6]),
                    sum(ind_monthly[6:9]),
                    sum(ind_monthly[9:12]),
                ]
            })

    # -------------------------
    # Project Totals (per project summary + per-indicator breakdown)
    # -------------------------
    project_totals = {}
    project_totals_chart = {}
    for project in projects:
        proj_qs = base_qs.filter(indicator__project=project)
        proj_rows = (
            proj_qs.values("indicator__id", "indicator__name", "indicator__unit")
            .annotate(total=Sum("value"))
            .order_by("indicator__name")
        )
        if proj_rows:
            project_totals[project.name] = list(proj_rows)
            project_totals_chart[project.name] = [
                {
                    "indicator": r["indicator__name"],
                    "unit": r["indicator__unit"],
                    "value": r["total"] or 0,
                }
                for r in proj_rows
            ]

    # ✅ Fix: aggregate per project (not empty anymore)
    project_totals_summary = (
        base_qs.values("indicator__project__id", "indicator__project__name")
        .annotate(total=Sum("value"))
        .order_by("indicator__project__name")
    )
    project_totals_labels = [p["indicator__project__name"] for p in project_totals_summary]
    project_totals_data = [p["total"] or 0 for p in project_totals_summary]

    # -------------------------
    # Project Trends
    # -------------------------
    project_trends = {}
    for project in projects:
        proj_qs = base_qs.filter(indicator__project=project)
        trend_data = []
        for ind in indicators.filter(project=project):
            ind_series = [
                proj_qs.filter(indicator=ind, month=m).aggregate(total=Sum("value"))["total"] or 0
                for m in monthly_labels
            ]
            if any(ind_series):
                trend_data.append({
                    "indicator": ind.name,
                    "unit": ind.unit,
                    "months": monthly_labels,
                    "values": ind_series,
                })
        if trend_data:
            project_trends[project.name] = trend_data

    context = {
        "year": year,
        "years": years,   # ✅ Updated: always includes current year
        "projects": projects,
        "indicators": indicators,
        "units": units,
        "selected_unit": selected_unit,
        "selected_indicators": selected_indicator_ids,

        "indicator_totals": indicator_totals,
        "selected_indicators_data": selected_indicators_data,
        "overall_total": overall_total,
        "ytd_totals": indicator_totals,

        "monthly_labels": monthly_labels,
        "monthly_data": monthly_data,
        "quarterly_labels": quarterly_labels,
        "quarterly_data": quarterly_data,

        "project_totals": project_totals,
        "project_totals_chart": project_totals_chart,
        "project_totals_labels": project_totals_labels,
        "project_totals_data": project_totals_data,
        "project_trends": project_trends,
    }
    return render(request, "core/reports.html", context)



# -----------------------------
# EXPORT PROJECT KPIs TO CSV
# -----------------------------
@login_required
def export_project_kpis(request, pk):
    """
    Export KPI monthly data for a given project to CSV.
    Only active indicators and active project are considered.
    """
    # Get project and ensure it is active
    project = Project.objects.filter(pk=pk, is_active=True).first()
    if not project:
        return HttpResponse("Project not found or inactive.", status=404)

    year = int(request.GET.get("year", datetime.date.today().year))  # fallback to current year

    # Prepare HTTP response for CSV
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{project.name}_kpis_{year}.csv"'

    writer = csv.writer(response)
    months = range(1, 13)
    header = ["Indicator", "Unit"] + [f"{m:02d}" for m in months] + ["Total"]
    writer.writerow(header)

    # Only active indicators
    indicators = project.indicators.filter(is_active=True)
    for ind in indicators:
        row = [ind.name, ind.unit]
        total = 0
        for m in months:
            entry = MonthlyEntry.objects.filter(
                indicator=ind,
                year=year
            ).filter(indicator__is_active=True, indicator__project__is_active=True).first()
            val = entry.value if entry else 0
            row.append(val)
            total += val
        row.append(total)
        writer.writerow(row)

    return response



# -----------------------------
# EXPORT PROJECT KPIs TO EXCEL (ACTIVE ONLY)
# -----------------------------

@login_required
def export_project_kpis_excel(request, pk):
    """
    Export KPI monthly data for a given project to Excel.
    Only active project and active indicators are included.
    """

    # Ensure project is active
    project = Project.objects.filter(pk=pk, is_active=True).first()
    if not project:
        return HttpResponse("Project not found or inactive.", status=404)

    year = int(request.GET.get("year", datetime.date.today().year))
    months = list(range(1, 13))
    month_headers = [calendar.month_abbr[m] for m in months]

    # Only active indicators
    indicators = project.indicators.filter(is_active=True)

    # Prepare target dictionary for active indicators
    target_dict = {
        ind.id: ind.targets.filter(year=year).first().value if ind.targets.filter(year=year).exists() else 0
        for ind in indicators
    }

    # Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{project.name} KPIs"

    # --- Title row ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + len(months))
    ws["A1"] = f"{project.name} - KPI Report {year}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # --- Header row ---
    headers = ["Annual Target", "Indicator", "Unit"] + month_headers + ["Total"]
    ws.append(headers)
    header_row = ws[ws.max_row]

    thick_border = Side(border_style="medium", color="000000")
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=thick_border, bottom=thick_border, left=thick_border, right=thick_border)

    # --- Data rows ---
    for ind in indicators:
        target = target_dict.get(ind.id, 0)
        row_values = []
        total = 0
        for m in months:
            # Filter by indicator, year, and month
            entry = MonthlyEntry.objects.filter(
                indicator=ind,
                year=year,
                month=m
            ).first()
            val = entry.value if entry else 0
            row_values.append(val)
            total += val

        excel_row = [target, ind.name, ind.unit] + row_values + [total]
        ws.append(excel_row)

    # --- Totals row ---
    monthly_totals = []
    for m in months:
        month_sum = sum(
            MonthlyEntry.objects.filter(
                indicator__in=indicators,
                year=year,
                month=m
            ).values_list('value', flat=True) or [0]
        )
        monthly_totals.append(month_sum)
    grand_total = sum(monthly_totals)
    totals_row = ["", "Total", ""] + monthly_totals + [grand_total]
    ws.append(totals_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # --- Footer note ---
    ws.merge_cells(start_row=ws.max_row + 2, start_column=1, end_row=ws.max_row + 2, end_column=len(headers))
    ws.cell(row=ws.max_row, column=1).value = f"Exported from Life Concern Data Management System - {datetime.date.today():%Y-%m-%d}"
    ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=10)

    # --- Borders and column widths ---
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 2):
        for cell in row:
            if cell.row == 2:
                continue
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if isinstance(cell.value, int):
                cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

    # --- Serve Excel file ---
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{project.name}_kpis_{year}.xlsx"'
    wb.save(response)
    return response



from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
import datetime
from .models import Indicator, MonthlyEntry, Project

@login_required
def monthly_trend_report(request):
    """
    AJAX view to return Monthly Trend data as JSON for chart rendering.
    Only includes active projects and active indicators.
    """
    current_year = datetime.date.today().year

    # --- Get filters from request ---
    try:
        selected_year = int(request.GET.get("year", current_year))
    except ValueError:
        selected_year = current_year

    selected_unit = request.GET.get("unit") or None
    selected_project = request.GET.get("project")
    selected_project = int(selected_project) if selected_project and selected_project.isdigit() else None
    selected_indicator = request.GET.get("indicator")
    selected_indicator = int(selected_indicator) if selected_indicator and selected_indicator.isdigit() else None

    # --- Filter indicators (only active) ---
    indicators = Indicator.objects.filter(is_active=True, project__is_active=True)
    if selected_unit:
        indicators = indicators.filter(unit=selected_unit)
    if selected_project:
        indicators = indicators.filter(project_id=selected_project)
    if selected_indicator:
        indicators = indicators.filter(id=selected_indicator)

    months = list(range(1, 13))
    month_labels = [datetime.date(1900, m, 1).strftime('%b') for m in months]

    # --- Prepare Monthly Trend data ---
    monthly_trends_data = []
    for ind in indicators:
        values = [
            int(
                MonthlyEntry.objects.filter(
                    year=selected_year,
                    month=m,
                    indicator=ind
                ).aggregate(total=Sum("value"))["total"] or 0
            )
            for m in months
        ]
        if any(values):
            monthly_trends_data.append({
                "indicator": ind.name,
                "project": ind.project.name,
                "values": values
            })

    return JsonResponse({
        "month_headers": month_labels,
        "monthly_trends_data": monthly_trends_data
    })


# -----------------------------
# More Reports View
# -----------------------------
@login_required
def more_reports(request):
    """
    Full HTML view that renders More Reports page with pivot table and charts.
    Only includes active projects, indicators, districts, and facilities.
    """

    import datetime
    from django.db.models import Q

    current_year = datetime.date.today().year

    # --- Selected filters from GET ---
    selected_year = request.GET.get("year")
    try:
        selected_year = int(selected_year)
    except (TypeError, ValueError):
        selected_year = current_year

    selected_unit = request.GET.get("unit") or None

    selected_project = request.GET.get("project")
    selected_project = int(selected_project) if selected_project and selected_project.isdigit() else None

    selected_indicator = request.GET.get("indicator")
    selected_indicator = int(selected_indicator) if selected_indicator and selected_indicator.isdigit() else None

    selected_district = request.GET.get("district") or None
    selected_facility = request.GET.get("facility") or None

    # --- Objects for selected filters ---
    selected_project_obj = Project.objects.filter(id=selected_project, is_active=True).first() if selected_project else None
    selected_indicator_obj = Indicator.objects.filter(
        id=selected_indicator,
        is_active=True,
        project__is_active=True
    ).first() if selected_indicator else None

    selected_district_obj = District.objects.filter(
        name=selected_district,
        is_active=True
    ).first() if selected_district else None

    selected_facility_obj = Facility.objects.filter(
        name=selected_facility,
        is_active=True
    ).first() if selected_facility else None

    # --- Available filter lists (only active) ---
    projects = Project.objects.filter(is_active=True).order_by("name")
    indicators = Indicator.objects.filter(is_active=True, project__is_active=True).order_by("name")
    districts = District.objects.filter(is_active=True).order_by("name")
    facilities = Facility.objects.filter(is_active=True).order_by("name")
    years = list(MonthlyEntry.objects.values_list("year", flat=True).distinct().order_by("year"))
    units = list(Indicator.objects.values_list("unit", flat=True).distinct())

    # --- IDs to check if selected_obj is already in the list ---
    project_ids = [p.id for p in projects]
    indicator_ids = [i.id for i in indicators]

    # --- Chart / pivot data (pass filters including district & facility) ---
    context = _get_report_data(
        selected_year=selected_year,
        selected_unit=selected_unit,
        selected_project=selected_project,
        selected_indicator=selected_indicator,
        selected_district=selected_district,
        selected_facility=selected_facility
    )

    # --- Add filter context variables for template ---
    context.update({
        "selected_year": selected_year,
        "selected_unit": selected_unit,
        "selected_project": selected_project,
        "selected_indicator": selected_indicator,
        "selected_district": selected_district,
        "selected_facility": selected_facility,
        "selected_project_obj": selected_project_obj,
        "selected_indicator_obj": selected_indicator_obj,
        "selected_district_obj": selected_district_obj,
        "selected_facility_obj": selected_facility_obj,
        "years": years,
        "units": units,
        "projects": projects,
        "indicators": indicators,
        "districts": districts,
        "facilities": facilities,
        "project_ids": project_ids,
        "indicator_ids": indicator_ids,
    })

    return render(request, "core/more_reports.html", context)




# -----------------------------
# Pivot Filter Data JSON (AJAX)
# -----------------------------
@login_required
def pivot_filter_data_json(request):
    """
    AJAX endpoint: returns filtered chart & heatmap data as JSON.
    Only includes active projects and active indicators.
    Respects district and facility filters.
    """
    import datetime

    current_year = datetime.date.today().year

    # --- Selected filters ---
    selected_year = request.GET.get("year")
    try:
        selected_year = int(selected_year)
    except (TypeError, ValueError):
        selected_year = current_year

    selected_unit = request.GET.get("unit") or None

    selected_project = request.GET.get("project")
    selected_project = int(selected_project) if selected_project and selected_project.isdigit() else None

    selected_indicator = request.GET.get("indicator")
    selected_indicator = int(selected_indicator) if selected_indicator and selected_indicator.isdigit() else None

    selected_district = request.GET.get("district") or None
    selected_facility = request.GET.get("facility") or None

    # --- Ensure selected objects are active ---
    selected_project_obj = Project.objects.filter(id=selected_project, is_active=True).first() if selected_project else None
    selected_indicator_obj = Indicator.objects.filter(
        id=selected_indicator,
        is_active=True,
        project__is_active=True
    ).first() if selected_indicator else None

    # --- Generate report data (active filtered + district/facility) ---
    context = _get_report_data(
        selected_year=selected_year,
        selected_unit=selected_unit,
        selected_project=selected_project_obj.id if selected_project_obj else None,
        selected_indicator=selected_indicator_obj.id if selected_indicator_obj else None,
        selected_district=selected_district,
        selected_facility=selected_facility
    )

    # --- Prepare JSON response ---
    data = {
        "indicator_distribution_labels": json.loads(context["indicator_distribution_labels"]),
        "indicator_distribution_data": json.loads(context["indicator_distribution_data"]),
        "indicator_progress_labels": json.loads(context["indicator_progress_labels"]),
        "indicator_progress_data": json.loads(context["indicator_progress_data"]),
        "cumulative_performance": json.loads(context["cumulative_performance"]),
        "top_indicators_labels": json.loads(context["top_indicators_labels"]),
        "top_indicators_data": json.loads(context["top_indicators_data"]),
        "facilities_json": json.loads(context["facilities_json"]),
        "correlation_data": json.loads(context["correlation_data"]),
        "yoy_labels": json.loads(context["yoy_labels"]),
        "yoy_data_values": json.loads(context["yoy_data_values"]),
        "month_headers": context["month_headers"],
    }

    return JsonResponse(data)

# -----------------------------
def _get_report_data(selected_year, selected_unit=None, selected_project=None, selected_indicator=None,
                     selected_district=None, selected_facility=None):
    """
    Helper function: computes all data for charts, heatmap, pivot table.
    Only active projects, active indicators, and active districts/facilities are included.
    """

    # --- Base queryset filtered by year and active status ---
    base_qs = MonthlyEntry.objects.filter(
        year=selected_year,
        indicator__is_active=True,
        indicator__project__is_active=True
    )

    # --- Apply unit/project/indicator filters ---
    if selected_unit:
        base_qs = base_qs.filter(indicator__unit=selected_unit)
    if selected_project:
        base_qs = base_qs.filter(indicator__project_id=selected_project)
    if selected_indicator:
        base_qs = base_qs.filter(indicator_id=selected_indicator)

    # --- Apply district/facility filters through related projects ---
    if selected_district or selected_facility:
        projects_qs = Project.objects.filter(is_active=True)

        if selected_district:
            projects_qs = projects_qs.filter(facilities__district__name=selected_district)
        if selected_facility:
            projects_qs = projects_qs.filter(facilities__name=selected_facility)

        project_ids = projects_qs.values_list('id', flat=True)
        base_qs = base_qs.filter(indicator__project_id__in=project_ids)

    # --- Indicators filtered by active projects/indicators ---
    indicators = Indicator.objects.filter(
        is_active=True,
        project__is_active=True
    )
    if selected_unit:
        indicators = indicators.filter(unit=selected_unit)
    if selected_project:
        indicators = indicators.filter(project_id=selected_project)
    if selected_indicator:
        indicators = indicators.filter(id=selected_indicator)
    if selected_district or selected_facility:
        indicators = indicators.filter(project_id__in=project_ids)

    # --- Pivot Table ---
    months = list(range(1, 13))
    month_headers = [calendar.month_abbr[m] for m in months]
    pivot_data = []
    column_totals = {m: 0 for m in months}
    grand_total = 0

    for ind in indicators:
        row_values = []
        row_total = 0
        for m in months:
            val = base_qs.filter(indicator=ind, month=m).aggregate(total=Sum("value"))["total"] or 0
            val = int(val)
            row_values.append(val)
            row_total += val
            column_totals[m] += val
        grand_total += row_total

        target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first()
        target_value = int(target_obj.value) if target_obj and target_obj.value > 0 else 1
        progress = row_total * 100 // target_value

        pivot_data.append({
            "year": selected_year,
            "project": ind.project.name if ind.project else "-",
            "indicator": ind.name,
            "unit": ind.unit,
            "values": row_values,
            "total": row_total,
            "target": target_value,
            "progress": progress,
        })

    pivot_data = sorted(pivot_data, key=lambda x: (x["project"], x["indicator"]))

    # --- Analytics ---
    indicator_distribution = base_qs.values("indicator__name").annotate(total=Sum("value")).order_by("-total")
    indicator_distribution_labels = [i["indicator__name"] for i in indicator_distribution]
    indicator_distribution_data = [int(i["total"]) for i in indicator_distribution]

    indicator_progress = []
    for ind in indicators:
        actual_total = int(base_qs.filter(indicator=ind).aggregate(total=Sum("value"))["total"] or 0)
        target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first()
        target_value = int(target_obj.value) if target_obj and target_obj.value > 0 else 1
        progress_pct = actual_total * 100 // target_value
        indicator_progress.append({"indicator": ind.name, "unit": ind.unit, "progress_pct": progress_pct})

    indicator_progress_labels = [i["indicator"] for i in indicator_progress]
    indicator_progress_data = [i["progress_pct"] for i in indicator_progress]

    cumulative_performance = []
    for ind in indicators:
        series = [int(base_qs.filter(indicator=ind, month=m).aggregate(total=Sum("value"))["total"] or 0) for m in months]
        running_total = 0
        cumulative_series = []
        for val in series:
            running_total += val
            cumulative_series.append(running_total)
        if any(series):
            cumulative_performance.append({"indicator": ind.name, "values": cumulative_series})

    totals = base_qs.values("indicator__name").annotate(total=Sum("value")).order_by("-total")
    top_indicators = list(totals[:5])
    top_indicators_labels = [i["indicator__name"] for i in top_indicators]
    top_indicators_data = [int(i["total"]) for i in top_indicators]

    # --- Heatmap (facilities linked to active projects) ---
    facilities_data = []
    facilities = Facility.objects.prefetch_related("projects").filter(is_active=True)
    if selected_district:
        facilities = facilities.filter(district__name=selected_district)
    if selected_facility:
        facilities = facilities.filter(name=selected_facility)

    for fac in facilities:
        projects_qs = fac.projects.filter(is_active=True)
        if selected_project:
            projects_qs = projects_qs.filter(id=selected_project)
        if selected_district or selected_facility:
            projects_qs = projects_qs.filter(id__in=project_ids)
        if not projects_qs.exists():
            continue

        fac_qs = base_qs.filter(indicator__project__in=projects_qs)
        if selected_indicator:
            fac_qs = fac_qs.filter(indicator_id=selected_indicator)
        if selected_unit:
            fac_qs = fac_qs.filter(indicator__unit=selected_unit)

        total_value = fac_qs.aggregate(total=Sum("value"))["total"] or 0
        if total_value == 0:
            continue

        facilities_data.append({
            "id": fac.id,
            "name": fac.name,
            "latitude": float(fac.latitude) if fac.latitude else None,
            "longitude": float(fac.longitude) if fac.longitude else None,
            "projects": [{"id": p.id, "name": p.name} for p in projects_qs],
            "total": int(total_value),
        })

    # --- Correlation ---
    correlation_data = []
    ind_list = list(indicators[:2])
    if len(ind_list) == 2:
        x_values = [int(base_qs.filter(indicator=ind_list[0], month=m).aggregate(total=Sum("value"))["total"] or 0) for m in months]
        y_values = [int(base_qs.filter(indicator=ind_list[1], month=m).aggregate(total=Sum("value"))["total"] or 0) for m in months]
        correlation_data = [{"x": x, "y": y} for x, y in zip(x_values, y_values)]
    elif len(ind_list) == 1:
        x_values = [int(base_qs.filter(indicator=ind_list[0], month=m).aggregate(total=Sum("value"))["total"] or 0) for m in months]
        correlation_data = [{"x": x, "y": 0} for x in x_values]

    # --- Year-over-Year ---
    available_years = MonthlyEntry.objects.values_list("year", flat=True).distinct().order_by("year")
    yoy_data_values = []
    yoy_labels = []
    for y in available_years:
        qs = MonthlyEntry.objects.filter(year=y, indicator__is_active=True, indicator__project__is_active=True)
        if selected_unit:
            qs = qs.filter(indicator__unit=selected_unit)
        if selected_project:
            qs = qs.filter(indicator__project_id=selected_project)
        if selected_indicator:
            qs = qs.filter(indicator_id=selected_indicator)
        if selected_district or selected_facility:
            qs = qs.filter(indicator__project_id__in=project_ids)
        total = int(qs.aggregate(total=Sum("value"))["total"] or 0)
        yoy_data_values.append(total)
        yoy_labels.append(y)

    return {
        "month_headers": month_headers,
        "indicator_distribution_labels": json.dumps(indicator_distribution_labels),
        "indicator_distribution_data": json.dumps(indicator_distribution_data),
        "indicator_progress_labels": json.dumps(indicator_progress_labels),
        "indicator_progress_data": json.dumps(indicator_progress_data),
        "cumulative_performance": json.dumps(cumulative_performance),
        "top_indicators_labels": json.dumps(top_indicators_labels),
        "top_indicators_data": json.dumps(top_indicators_data),
        "facilities_json": json.dumps(facilities_data),
        "correlation_data": json.dumps(correlation_data),
        "yoy_labels": json.dumps(yoy_labels),
        "yoy_data_values": json.dumps(yoy_data_values),
        "pivot_data": pivot_data,
        "column_totals": column_totals,
        "grand_total": grand_total,
    }




# ------------------------
# Helper: get short month name
# ------------------------
def get_month_name(month_number):
    return calendar.month_abbr[month_number] if 1 <= month_number <= 12 else ""

# ------------------------
# CSV export view
# ------------------------

def get_month_name(m):
    import calendar
    return calendar.month_abbr[m]

@login_required
def more_reports_export_csv(request):
    # ------------------------
    # Filters from GET
    # ------------------------
    selected_year = request.GET.get("year")
    selected_unit = request.GET.get("unit")
    selected_project = request.GET.get("project")
    selected_indicator = request.GET.get("indicator")

    # Cast numeric filters
    try:
        selected_year = int(selected_year)
    except (TypeError, ValueError):
        selected_year = None

    try:
        selected_project = int(selected_project)
    except (TypeError, ValueError):
        selected_project = None

    try:
        selected_indicator = int(selected_indicator)
    except (TypeError, ValueError):
        selected_indicator = None

    if selected_unit:
        selected_unit = selected_unit.strip()

    # ------------------------
    # Queryset of active indicators
    # ------------------------
    indicators = Indicator.objects.filter(
        is_active=True,
        project__is_active=True
    )

    if selected_project:
        indicators = indicators.filter(project_id=selected_project)
    if selected_unit:
        indicators = indicators.filter(unit=selected_unit)
    if selected_indicator:
        indicators = indicators.filter(id=selected_indicator)

    indicators = indicators.select_related("project").order_by("project__name", "name")

    # ------------------------
    # Get month headers
    # ------------------------
    months = list(range(1, 13))
    month_headers = [get_month_name(m) for m in months]

    # ------------------------
    # CSV response setup
    # ------------------------
    filename_parts = ["pivot"]
    if selected_year:
        filename_parts.append(str(selected_year))
    if selected_project:
        project_obj = Project.objects.filter(id=selected_project, is_active=True).first()
        if project_obj:
            filename_parts.append(project_obj.name.replace(" ", "_"))
    if selected_unit:
        filename_parts.append(selected_unit)
    filename = "_".join(filename_parts) + ".csv"

    response = HttpResponse(content_type="text/csv")
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)

    # ------------------------
    # Write header row
    # ------------------------
    headers = ["Year", "Project", "Indicator", "Unit"] + month_headers + ["Total", "Target", "Progress (%)"]
    writer.writerow(headers)

    # ------------------------
    # Totals tracking
    # ------------------------
    column_totals = {m: 0 for m in months}
    grand_total = 0

    # ------------------------
    # Write each indicator row
    # ------------------------
    for ind in indicators:
        row_values = []
        row_total = 0

        for m in months:
            qs = MonthlyEntry.objects.filter(
                indicator=ind,
                month=m,
                indicator__is_active=True,
                indicator__project__is_active=True
            )
            if selected_year:
                qs = qs.filter(year=selected_year)
            total_val = qs.aggregate(total=Sum("value"))["total"] or 0
            total_val = int(total_val)
            row_values.append(total_val)
            row_total += total_val
            column_totals[m] += total_val

        grand_total += row_total

        # Target and progress
        target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first() if selected_year else None
        target_value = int(target_obj.value) if target_obj and target_obj.value else getattr(ind, "target", 0) or 0
        progress = round((row_total / target_value * 100), 1) if target_value else 0

        writer.writerow([
            selected_year or "-",
            ind.project.name if ind.project else "-",
            ind.name,
            ind.unit,
            *row_values,
            row_total,
            target_value,
            f"{progress}%"
        ])

    # ------------------------
    # Write totals row
    # ------------------------
    totals_row = ["Column Totals", "", "", ""] + [column_totals[m] for m in months] + [grand_total, "-", "-"]
    writer.writerow(totals_row)

    return response





# -----------------------------
# Edit Indicator Targets View
# -----------------------------
@login_required
def edit_indicator_targets(request, indicator_id):
    indicator = get_object_or_404(Indicator, id=indicator_id)
    queryset = IndicatorTarget.objects.filter(indicator=indicator)

    # Create a modelformset with at least 1 extra blank form for new targets
    IndicatorTargetFormSetClass = modelformset_factory(
        IndicatorTarget,
        form=IndicatorTargetForm,
        extra=1,        # at least one blank form for adding
        can_delete=True # allows deletion
    )

    formset = IndicatorTargetFormSetClass(request.POST or None, queryset=queryset)
    numeric_errors = []

    if request.method == "POST" and formset.is_valid():
        try:
            with transaction.atomic():  # ensures atomic DB operations

                # 1️⃣ Delete targets marked for deletion
                for form in formset.deleted_forms:
                    if form.instance.pk:
                        form.instance.delete()

                # 2️⃣ Save or update remaining targets
                for form in formset:
                    if form.cleaned_data.get('DELETE'):
                        continue  # skip deleted forms

                    year = form.cleaned_data.get('year')
                    raw_value = form.cleaned_data.get('value')

                    if year is None or raw_value in [None, ""]:
                        continue  # skip incomplete forms

                    # Ensure value is a whole number
                    try:
                        value = safe_int(raw_value)
                    except (ValueError, TypeError):
                        numeric_errors.append(f"Year {year}: '{raw_value}' is not a whole number.")
                        continue

                    # Update existing target or create new one
                    IndicatorTarget.objects.update_or_create(
                        indicator=indicator,
                        year=year,
                        defaults={'value': value}
                    )

        except IntegrityError:
            formset.add_error(None, "Error: Duplicate target for the same year.")

        if numeric_errors:
            messages.warning(
                request,
                "Some targets were not saved due to invalid values: " + ", ".join(numeric_errors)
            )
        else:
            messages.success(request, "Indicator targets saved successfully.")

        return redirect('project_kpis', pk=indicator.project.id)

    # Prepare context with current year and year range
    current_year = datetime.date.today().year
    year_range = range(current_year - 5, current_year + 6)

    context = {
        "indicator": indicator,
        "formset": formset,
        "current_year": current_year,
        "year_range": year_range,
    }

    return render(request, "core/edit_indicator_targets.html", context)



@login_required
def data_story(request):
    today = date.today()
    current_year = today.year

    # -------------------------
    # Year Filter: last 5 years, default to current year
    # -------------------------
    all_years = MonthlyEntry.objects.dates("created_at", "year", order="DESC")
    available_years = sorted({y.year for y in all_years}, reverse=True)
    year_list = list(range(current_year, current_year - 5, -1))
    for y in available_years:
        if y not in year_list:
            year_list.append(y)
    try:
        selected_year = int(request.GET.get("year")) if request.GET.get("year") else current_year
    except (ValueError, TypeError):
        selected_year = current_year

    # -------------------------
    # Last completed month
    # -------------------------
    latest_month = today.month - 1 if today.month > 1 else 12
    latest_month_year = selected_year if today.month > 1 else selected_year - 1

    # -------------------------
    # Select active projects
    # -------------------------
    all_projects = Project.objects.filter(is_active=True).order_by("name")
    selected_projects = all_projects

    # -------------------------
    # Summary Counters
    # -------------------------
    total_kpis = 0
    improving_count = declining_count = stable_count = new_count = 0
    performance_score_sum = 0
    performance_score_items = 0
    project_insights = []
    missing_kpis = []

    # -------------------------
    # Iterate Projects + Indicators
    # -------------------------
    for project in selected_projects:
        indicators = Indicator.objects.filter(is_active=True, project=project).order_by("name")
        insights = []

        for indicator in indicators:
            entries = list(MonthlyEntry.objects.filter(indicator=indicator, year=selected_year).order_by("month"))
            month_labels = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
            monthly_actuals = [0.0] * 12
            for e in entries:
                monthly_actuals[e.month - 1] = float(e.value or 0)

            # -------------------------
            # Target
            # -------------------------
            target_obj = IndicatorTarget.objects.filter(indicator=indicator, year=selected_year).first()
            try:
                annual_target = float(target_obj.value) if target_obj else 0.0
            except (InvalidOperation, ValueError, TypeError):
                annual_target = 0.0
            monthly_target_value = round(annual_target / 12, 2) if annual_target else 0
            monthly_targets = [monthly_target_value] * 12

            # -------------------------
            # Latest entry = last completed month
            # -------------------------
            latest_entry = MonthlyEntry.objects.filter(
                indicator=indicator,
                year=selected_year,
                month=latest_month
            ).first()

            # Previous entry before latest month
            prev_entry = MonthlyEntry.objects.filter(
                indicator=indicator,
                created_at__lt=latest_entry.created_at if latest_entry else None
            ).order_by("-created_at").first() if latest_entry else None

            current_month_actual = float(latest_entry.value) if latest_entry else 0.0
            cumulative_actual = sum([float(e.value or 0) for e in entries])
            cumulative_target = monthly_target_value * (latest_entry.month if latest_entry else 12)

            # -------------------------
            # Accurate cumulative percentage (no cap at 100%)
            # -------------------------
            cumulative_pct = (cumulative_actual / cumulative_target * 100) if cumulative_target else 0
            cumulative_pct = round(cumulative_pct, 1)

            # -------------------------
            # Determine interpretation (target insight)
            # -------------------------
            if annual_target:
                if cumulative_pct < 50:
                    target_note = "far below target"
                elif 50 <= cumulative_pct < 90:
                    target_note = "progressing but below target"
                elif 90 <= cumulative_pct < 100:
                    target_note = "close to target"
                elif 100 <= cumulative_pct <= 110:
                    target_note = "target reached"
                else:
                    target_note = "exceeded target"
            else:
                target_note = "target unavailable"

            # -------------------------
            # Trend calculation
            # -------------------------
            change = ((current_month_actual - float(prev_entry.value)) / float(prev_entry.value) * 100
                      if prev_entry and prev_entry.value != 0 else 0.0)

            last_3_entries = entries[-3:] if len(entries) >= 3 else entries
            trend_vals = []
            for i in range(1, len(last_3_entries)):
                prev_val = float(last_3_entries[i-1].value or 0)
                curr_val = float(last_3_entries[i].value or 0)
                if curr_val > prev_val: trend_vals.append("up")
                elif curr_val < prev_val: trend_vals.append("down")
                else: trend_vals.append("stable")

            trend_summary = "stable"
            if all(t == "up" for t in trend_vals) and trend_vals:
                trend_summary = "upward"
            elif all(t == "down" for t in trend_vals) and trend_vals:
                trend_summary = "downward"

            # -------------------------
            # Insight story (more accurate + meaningful)
            # -------------------------
            if not entries:
                story = (f"No data recorded yet for this year. "
                         f"Cumulative target is {annual_target:.1f}, monthly target {monthly_target_value:.1f}. "
                         f"Trend cannot be determined yet.")
                tag = "New"
                new_count += 1
                missing_kpis.append(indicator.name)

            elif current_month_actual > 0 and not prev_entry:
                story = (f"Current month actual is {current_month_actual:.1f}, cumulative actual {cumulative_actual:.1f} "
                         f"({cumulative_pct:.1f}% of {annual_target:.1f} target). Data entry has just begun this period.")
                tag = "New"
                new_count += 1

            else:
                if change > 0:
                    tag = "Improving"; improving_count += 1
                    story = (
                        f"Current month actual {current_month_actual:.1f}, cumulative {cumulative_actual:.1f} "
                        f"({cumulative_pct:.1f}% of {annual_target:.1f}). "
                        f"Increased by {change:.1f}% from last month, showing a {trend_summary} trend — {target_note}."
                    )
                elif change < 0:
                    tag = "Declining"; declining_count += 1
                    story = (
                        f"Current month actual {current_month_actual:.1f}, cumulative {cumulative_actual:.1f} "
                        f"({cumulative_pct:.1f}% of {annual_target:.1f}). "
                        f"Decreased by {abs(change):.1f}% from last month, {trend_summary} trend — {target_note}."
                    )
                else:
                    tag = "Stable"; stable_count += 1
                    story = (
                        f"Actual {current_month_actual:.1f}, cumulative {cumulative_actual:.1f} "
                        f"({cumulative_pct:.1f}% of {annual_target:.1f}). "
                        f"Stable performance with {trend_summary} pattern — {target_note}."
                    )

            # -------------------------
            # Chart Data
            # -------------------------
            chart_data = json.dumps({
                "labels": month_labels,
                "actual": monthly_actuals,
                "target": monthly_targets,
            })

            # -------------------------
            # Performance aggregation
            # -------------------------
            if cumulative_target > 0:
                progress_percent = (cumulative_actual / cumulative_target * 100)
                performance_score_sum += min(progress_percent, 120)  # allow >100 for realism
                performance_score_items += 1

            total_kpis += 1

            insights.append({
                "indicator_id": indicator.id,
                "title": indicator.name,
                "chart_data": chart_data,
                "story": story,
                "tag": tag,
                "change": change,
                "actual": current_month_actual,
                "target": annual_target,
            })

        if insights:
            sorted_insights = sorted(insights, key=lambda x: x.get("change", 0), reverse=True)
            project_insights.append({
                "project": project,
                "best": sorted_insights[:2],
                "others": sorted_insights[2:-2] if len(sorted_insights) > 4 else sorted_insights[2:-2],
                "worst": sorted_insights[-2:] if len(sorted_insights) >= 2 else insights,
            })

    # -------------------------
    # Summary Percentages
    # -------------------------
    improving_pct = (improving_count / total_kpis * 100) if total_kpis else 0
    declining_pct = (declining_count / total_kpis * 100) if total_kpis else 0
    performance_score = (performance_score_sum / performance_score_items) if performance_score_items else 0

    # -------------------------
    # Dynamic overall narrative
    # -------------------------
    no_data_count = len(missing_kpis)
    perf = performance_score
    overall_story = (
        f"The organization is tracking {total_kpis} key performance indicators across active projects. "
        f"Overall performance stands at {perf:.1f}%, indicating a "
        f"{'strong' if perf > 85 else 'steady' if perf > 60 else 'moderate' if perf > 40 else 'weak'} achievement rate. "
        f"A total of {improving_pct:.1f}% of indicators are improving, while {declining_pct:.1f}% are declining, "
        f"showing that performance momentum is {'positive' if improving_pct > declining_pct else 'mixed' if abs(improving_pct - declining_pct) < 10 else 'concerning'}. "
        f"{'Some indicators have surpassed their targets, showing outstanding progress. ' if perf > 100 else ''}"
        f"Indicators without recent data ({no_data_count}) such as {', '.join(missing_kpis[:5]) if missing_kpis else 'none'} "
        f"limit accurate tracking and may underestimate performance. "
        f"Efforts should focus on data completeness and underperforming areas to maintain gains. "
        f"The trend reflects {'steady improvement' if improving_pct > 50 else 'areas needing renewed focus'}, "
        f"with targets {'mostly met' if perf > 85 else 'partially achieved' if perf > 60 else 'largely unmet'}. "
        f"Maintaining consistent reporting will ensure an accurate reflection of progress."
    )

    # -------------------------
    # Context
    # -------------------------
    context = {
        "project_insights": project_insights,
        "selected_year": selected_year,
        "year_list": year_list,
        "total_kpis": total_kpis,
        "percent_improving": f"{improving_pct:.1f}%",
        "percent_declining": f"{declining_pct:.1f}%",
        "performance_score": f"{performance_score:.1f}%",
        "overall_story": overall_story,
    }

    return render(request, "core/data_story.html", context)


# -------------------------
# EXPORT CSV - Dynamic with Numeric Totals
# -------------------------

@login_required
def reports_export_csv(request):
    # Fetch all indicators
    indicators = Indicator.objects.all()

    # Dynamically determine fields to export (exclude ID if desired)
    exclude_fields = ['id']
    fields = [f.name for f in Indicator._meta.get_fields() if f.concrete and f.name not in exclude_fields]

    # Identify numeric fields for totals
    numeric_fields = [f.name for f in Indicator._meta.get_fields()
                      if f.concrete and isinstance(f, FloatField)]

    # Initialize totals dictionary
    totals = {f: 0 for f in numeric_fields}

    # Prepare CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="report.csv"'

    writer = csv.writer(response)

    # Write header
    writer.writerow(fields)

    # Write data rows and accumulate totals
    for ind in indicators:
        row = []
        for f in fields:
            val = getattr(ind, f, "")
            row.append(val)
            # accumulate total if numeric
            if f in numeric_fields and isinstance(val, (int, float)):
                totals[f] += val
        writer.writerow(row)

    # Write totals row
    totals_row = []
    for f in fields:
        if f in numeric_fields:
            totals_row.append(totals[f])
        else:
            totals_row.append("Total" if f == fields[0] else "")
    writer.writerow(totals_row)

    return response



@login_required
def more_reports_export_excel(request):
    # --- Selected filters ---
    selected_year = request.GET.get("year")
    try:
        selected_year = int(selected_year)
    except (TypeError, ValueError):
        selected_year = datetime.date.today().year

    selected_unit = request.GET.get("unit") or None
    selected_project = request.GET.get("project")
    selected_project = int(selected_project) if selected_project and selected_project.isdigit() else None
    selected_indicator = request.GET.get("indicator")
    selected_indicator = int(selected_indicator) if selected_indicator and selected_indicator.isdigit() else None

    # --- Base queryset ---
    base_qs = MonthlyEntry.objects.filter(year=selected_year)
    if selected_unit:
        base_qs = base_qs.filter(indicator__unit=selected_unit)
    if selected_project:
        base_qs = base_qs.filter(indicator__project_id=selected_project)
    if selected_indicator:
        base_qs = base_qs.filter(indicator_id=selected_indicator)

    # --- Indicators (active only) ---
    indicators = Indicator.objects.filter(is_active=True)
    if selected_unit:
        indicators = indicators.filter(unit=selected_unit)
    if selected_project:
        indicators = indicators.filter(project_id=selected_project)
    if selected_indicator:
        indicators = indicators.filter(id=selected_indicator)

    # --- Pivot Table Data ---
    months = list(range(1, 13))
    month_headers = [calendar.month_abbr[m] for m in months]
    pivot_data = []
    column_totals = {m: 0 for m in months}
    grand_total = 0

    for ind in indicators:
        row_values, row_total = [], 0
        for m in months:
            val = base_qs.filter(indicator=ind, month=m).aggregate(total=Sum("value"))["total"] or 0
            val = int(val)
            row_values.append(val)
            row_total += val
            column_totals[m] += val
        grand_total += row_total

        target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first()
        target_value = int(target_obj.value) if target_obj and target_obj.value > 0 else 1
        progress = row_total / target_value  # decimal for % formatting

        pivot_data.append({
            "year": selected_year,
            "project": ind.project.name if ind.project else "-",
            "indicator": ind.name,
            "unit": ind.unit,
            "values": row_values,
            "total": row_total,
            "target": target_value,
            "progress": progress,
        })

    # Sort pivot_data
    pivot_data = sorted(pivot_data, key=lambda x: (x["project"], x["indicator"]))

    # --- Create Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Pivot Table"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(months) + 3)
    ws["A1"] = "Life Concern - Pivot Table showing Projects, Indicators and Progress"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Year", "Project", "Indicator", "Unit"] + month_headers + ["Total", "Target", "Progress (%)"]
    ws.append(headers)
    header_row = ws[ws.max_row]

    # Header borders
    thick_border = Side(border_style="medium", color="000000")
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=thick_border, bottom=thick_border)

    # Data rows
    for row in pivot_data:
        excel_row = [
            row["year"], row["project"], row["indicator"], row["unit"],
            *row["values"], row["total"], row["target"], row["progress"]
        ]
        ws.append(excel_row)

        # Format progress
        progress_cell = ws.cell(row=ws.max_row, column=len(excel_row))
        progress_cell.number_format = "0%"
        progress_val = row["progress"] * 100

        if progress_val >= 75:
            progress_cell.font = Font(bold=True, color="006100")  # dark green
        elif progress_val >= 50:
            progress_cell.font = Font(bold=True, color="9C6500")  # dark amber
        else:
            progress_cell.font = Font(bold=True, color="9C0006")  # dark red
        progress_cell.alignment = Alignment(horizontal="center")

    # Totals row
    totals_row = ["", "", "", "Column Totals"] + [column_totals[m] for m in months] + [grand_total, "-", "-"]
    ws.append(totals_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    # Footer note
    ws.merge_cells(start_row=ws.max_row + 2, start_column=1, end_row=ws.max_row + 2, end_column=len(headers))
    ws.cell(row=ws.max_row, column=1).value = f"Exported from Life Concern Data Management System - {datetime.date.today():%Y-%m-%d}"
    ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=10)

    # Apply subtle borders
    thin = Side(border_style="thin", color="000000")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            if cell.row == 2:
                continue
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Auto column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 3

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="Pivot_Report_{selected_year}.xlsx"'
    wb.save(response)
    return response


@login_required
def monthly_performance_report(request):
    # --- Year filter ---
    current_year = datetime.date.today().year
    selected_year = request.GET.get("year")
    try:
        selected_year = int(selected_year)
    except (TypeError, ValueError):
        selected_year = current_year

    # --- Unit & Indicator filters ---
    selected_unit = request.GET.get("unit") or None
    selected_indicator = request.GET.get("indicator")
    selected_indicator = int(selected_indicator) if selected_indicator and selected_indicator.isdigit() else None

    # --- Month range filter ---
    month_from = request.GET.get("month_from")
    month_to = request.GET.get("month_to")

    try:
        month_from = int(month_from) if month_from else 1
        month_to = int(month_to) if month_to else 12
    except ValueError:
        month_from = 1
        month_to = 12

    # Ensure valid range
    if month_from < 1 or month_from > 12:
        month_from = 1
    if month_to < 1 or month_to > 12:
        month_to = 12
    if month_from > month_to:
        month_from, month_to = month_to, month_from

    # --- Base queryset for monthly entries ---
    base_qs = MonthlyEntry.objects.filter(year=selected_year)
    if selected_unit:
        base_qs = base_qs.filter(indicator__unit=selected_unit)
    if selected_indicator:
        base_qs = base_qs.filter(indicator_id=selected_indicator)

    # --- Months and headers based on selected range ---
    months = list(range(month_from, month_to + 1))
    month_headers = [calendar.month_abbr[m] for m in months]

    # --- Projects (only active) ---
    projects = Project.objects.filter(is_active=True).order_by("name")

    # --- Monthly trends per project ---
    monthly_trends = []
    for proj in projects:
        proj_qs = base_qs.filter(indicator__project=proj)
        series = [
            int(proj_qs.filter(month=m).aggregate(total=Sum("value"))["total"] or 0)
            for m in months
        ]

        # Keep project even if all values = 0
        indicators_list = sorted(proj_qs.values_list("indicator__name", flat=True).distinct())
        monthly_trends.append({
            "project_id": proj.id,
            "project_name": proj.name,
            "indicator_summary": ", ".join(indicators_list) if indicators_list else "No indicators",
            "values": series,
        })

    # --- Units & indicators for filter dropdowns ---
    units = sorted(Indicator.objects.values_list("unit", flat=True).distinct())
    indicators = Indicator.objects.all().order_by("name")

    # --- Years for filter dropdown ---
    years = (
        list(MonthlyEntry.objects.values_list("year", flat=True).distinct().order_by("-year"))
        or [current_year]
    )

    # --- Month choices for dropdowns ---
    month_choices = [(i, calendar.month_abbr[i]) for i in range(1, 13)]

    context = {
        "selected_year": selected_year,
        "years": years,
        "selected_unit": selected_unit,
        "selected_indicator": selected_indicator,
        "units": units,
        "indicators": indicators,
        "months": months,
        "month_headers": month_headers,
        "monthly_trends": monthly_trends,
        "month_from": month_from,
        "month_to": month_to,
        "month_choices": month_choices,
    }

    return render(request, "core/monthly_performance_report.html", context)



from django import forms
from django.conf import settings
from .models import Facility, Project, District
import json, os


# -----------------------------
# Facility Form
# -----------------------------
class FacilityForm(forms.ModelForm):
    class Meta:
        model = Facility
        fields = ['name', 'latitude', 'longitude']
        widgets = {
            'name': forms.TextInput(attrs={'class': 'form-control'}),
            'latitude': forms.NumberInput(attrs={'class': 'form-control', 'step': '0.000001'}),
            'longitude': forms.NumberInput(attrs={'class': 'form-control', 'step': '0.000001'}),
        }


# -----------------------------
# Detect district from lat/lon
# -----------------------------
def get_district_for_location(lat, lon):
    districts = District.objects.all()
    # Try bounding box first
    district = District.objects.filter(
        min_latitude__lte=lat,
        max_latitude__gte=lat,
        min_longitude__lte=lon,
        max_longitude__gte=lon
    ).first()
    if district:
        return district

    # Fallback: nearest by center coordinates
    def distance(d):
        if d.latitude is None or d.longitude is None:
            return float('inf')
        return (float(d.latitude) - lat) ** 2 + (float(d.longitude) - lon) ** 2

    return min(districts, key=distance) if districts.exists() else None


# -----------------------------
# Facilities Map View
# -----------------------------
def facilities_map(request):
    active_projects = Project.objects.filter(is_active=True)
    current_year = date.today().year
    year = int(request.GET.get("year", current_year))
    year_options = list(range(current_year - 5, current_year + 6))

    # Handle POST (Add/Edit)
    if request.method == "POST":
        facility_id = request.POST.get("facility_id")
        project_ids = request.POST.getlist("projects")

        if facility_id:
            facility = get_object_or_404(Facility, id=facility_id)
            form = FacilityForm(request.POST, instance=facility)
        else:
            form = FacilityForm(request.POST)

        if form.is_valid():
            facility = form.save(commit=False)
            district = get_district_for_location(float(facility.latitude), float(facility.longitude))
            facility.district = district
            facility.save()

            # Update projects
            if project_ids:
                facility_projects = Project.objects.filter(id__in=project_ids)
                facility.projects.set(facility_projects)
            else:
                facility.projects.clear()

            # Return updated map data
            return _map_data_response(active_projects, year)

        return JsonResponse({"success": False, "errors": form.errors})

    # -------------------------
    # GET request
    # -------------------------
    facilities_list = _get_facilities_list(active_projects, year)
    active_districts = _get_active_districts(active_projects)

    # ✅ Load GeoJSON (locally cached) for shading
    geojson_path = os.path.join(settings.BASE_DIR, "static", "geojson", "geojsonmalawi_districts.geojson")
    if os.path.exists(geojson_path):
        with open(geojson_path, "r", encoding="utf-8") as f:
            districts_geojson = json.load(f)
    else:
        districts_geojson = {}

    return render(request, "core/facilities_map.html", {
        "form": FacilityForm(),
        "facilities_json": json.dumps(facilities_list),
        "active_projects": active_projects,
        "year_options": year_options,
        "selected_year": year,
        "districts_geojson": json.dumps(districts_geojson),
        "active_districts": json.dumps(active_districts),
    })


# -----------------------------
# Helper: Facility + Districts
# -----------------------------
def _get_facilities_list(active_projects, year):
    facilities_qs = (
        Facility.objects
        .filter(projects__in=active_projects)
        .distinct()
        .select_related('district')
        .prefetch_related('projects')
    )

    facilities_list = []
    for f in facilities_qs:
        projects_data = []
        for p in f.projects.all():
            people_reached = sum(
                ind.entries.filter(year=year).aggregate(total=models.Sum('value'))['total'] or 0
                for ind in p.indicators.filter(unit__iexact='people reached')
            )
            projects_data.append({
                "id": p.id,
                "name": p.name,
                "people_reached": people_reached,
            })
        facilities_list.append({
            "id": f.id,
            "name": f.name,
            "latitude": float(f.latitude),
            "longitude": float(f.longitude),
            "district": f.district.name if f.district else None,
            "projects": projects_data,
        })
    return facilities_list


def _get_active_districts(active_projects):
    return list(
        Facility.objects
        .filter(projects__in=active_projects)
        .exclude(district=None)
        .values_list('district__name', flat=True)
        .distinct()
    )


def _map_data_response(active_projects, year):
    facilities_list = _get_facilities_list(active_projects, year)
    active_districts = _get_active_districts(active_projects)
    return JsonResponse({
        "success": True,
        "facilities": facilities_list,
        "active_districts": active_districts,
    })


# -----------------------------
# Delete Facility
# -----------------------------
def facility_delete(request, pk):
    facility = get_object_or_404(Facility, pk=pk)
    facility.delete()
    return JsonResponse({"success": True})


from weasyprint import HTML

@login_required
def more_reports_pdf(request):
    current_year = datetime.date.today().year

    # --- Selected filters ---
    selected_year = request.GET.get("year")
    try:
        selected_year = int(selected_year)
    except (TypeError, ValueError):
        selected_year = current_year

    selected_unit = request.GET.get("unit") or None
    selected_project = request.GET.get("project")
    selected_project = int(selected_project) if selected_project and selected_project.isdigit() else None
    selected_indicator = request.GET.get("indicator")
    selected_indicator = int(selected_indicator) if selected_indicator and selected_indicator.isdigit() else None

    # ✅ New: Active/Inactive filter
    selected_status = request.GET.get("status")  # expected values: "active", "inactive", or None

    # --- Base queryset ---
    base_qs = MonthlyEntry.objects.filter(year=selected_year)
    if selected_unit:
        base_qs = base_qs.filter(indicator__unit=selected_unit)
    if selected_project:
        base_qs = base_qs.filter(indicator__project_id=selected_project)
    if selected_indicator:
        base_qs = base_qs.filter(indicator_id=selected_indicator)
    if selected_status == "active":
        base_qs = base_qs.filter(indicator__is_active=True)
    elif selected_status == "inactive":
        base_qs = base_qs.filter(indicator__is_active=False)

    # --- Indicators ---
    indicators = Indicator.objects.all()
    if selected_unit:
        indicators = indicators.filter(unit=selected_unit)
    if selected_project:
        indicators = indicators.filter(project_id=selected_project)
    if selected_indicator:
        indicators = indicators.filter(id=selected_indicator)
    if selected_status == "active":
        indicators = indicators.filter(is_active=True)
    elif selected_status == "inactive":
        indicators = indicators.filter(is_active=False)

    # Sort indicators by project and name
    indicators = indicators.order_by('project__name', 'name')

    # --- Pivot Table Data ---
    months = list(range(1, 13))
    month_headers = [calendar.month_abbr[m] for m in months]
    pivot_data = []
    column_totals = {m: 0 for m in months}
    grand_total = 0

    for ind in indicators:
        row_values, row_total = [], 0
        for m in months:
            val = base_qs.filter(indicator=ind, month=m).aggregate(total=Sum("value"))["total"] or 0
            val = int(val)
            row_values.append(val)
            row_total += val
            column_totals[m] += val
        grand_total += row_total

        target_obj = IndicatorTarget.objects.filter(indicator=ind, year=selected_year).first()
        target_value = int(target_obj.value) if target_obj and target_obj.value > 0 else 1
        progress = (row_total / target_value) * 100  # percentage

        # Determine CSS class for progress
        if progress >= 75:
            progress_class = "progress-high"
        elif progress >= 50:
            progress_class = "progress-medium"
        else:
            progress_class = "progress-low"

        pivot_data.append({
            "year": selected_year,
            "project": ind.project.name if ind.project else "-",
            "indicator": ind.name,
            "unit": ind.unit,
            "status": "Active" if getattr(ind, "is_active", True) else "Inactive",  # ✅ Added status field
            "values": row_values,
            "total": row_total,
            "target": target_value,
            "progress": f"{progress:.1f}%",
            "progress_class": progress_class
        })

    # --- Render HTML ---
    html_string = render_to_string('core/more_reports_pdf.html', {
        "month_headers": month_headers,
        "pivot_data": pivot_data,
        "column_totals": column_totals,
        "grand_total": grand_total,
        "now": datetime.date.today(),
        "selected_status": selected_status,  # ✅ Pass to template if needed
    })

    # --- Generate PDF ---
    pdf_file = HTML(string=html_string).write_pdf()

    # --- Return PDF response ---
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Pivot_Report_{selected_year}.pdf"'
    return response



from io import BytesIO
from weasyprint import HTML
from openpyxl.styles import Font, Alignment, Border, Side
from .models import MonthlyEntry, Project, Indicator


def actual_vs_variance_report(request):
    today = datetime.date.today()
    current_year = today.year

    # --- Filters ---
    year = int(request.GET.get('year', current_year))
    month_from = int(request.GET.get('month_from', 1))
    month_to = int(request.GET.get('month_to', 12))
    indicator_id = request.GET.get('indicator')
    unit = request.GET.get('unit')
    project_id = request.GET.get('project')
    status = request.GET.get('status')  # active / inactive / None

    # --- Month choices ---
    month_choices = [(i, calendar.month_abbr[i]) for i in range(1, 13)]

    # --- Determine default active filtering ---
    if status not in ("active", "inactive"):
        status = "active"

    # --- Projects queryset ---
    projects_qs = Project.objects.all()
    if status == "active":
        projects_qs = projects_qs.filter(is_active=True)
    elif status == "inactive":
        projects_qs = projects_qs.filter(is_active=False)

    # --- Indicators queryset ---
    indicators_qs = Indicator.objects.all()
    if status == "active":
        indicators_qs = indicators_qs.filter(is_active=True, project__is_active=True)
    elif status == "inactive":
        indicators_qs = indicators_qs.filter(is_active=False, project__is_active=False)

    if indicator_id:
        indicators_qs = indicators_qs.filter(id=indicator_id)
    if project_id:
        indicators_qs = indicators_qs.filter(project_id=project_id)
    if unit:
        indicators_qs = indicators_qs.filter(unit=unit)

    # --- Entries queryset ---
    entries = MonthlyEntry.objects.filter(year=year, month__gte=month_from, month__lte=month_to)
    entries = entries.filter(indicator__in=indicators_qs)

    # --- Build table rows ---
    table_rows = []
    for indicator in indicators_qs:
        target_obj = indicator.targets.filter(year=year).first()
        annual_target = target_obj.value if target_obj else 0
        monthly_target = round(annual_target / 12, 2) if annual_target else 0

        indicator_entries = entries.filter(indicator=indicator)
        actuals_list = [0] * 12
        for entry in indicator_entries:
            actuals_list[entry.month - 1] = entry.value or 0

        cumulative_actuals = sum(actuals_list[month_from - 1:month_to])
        current_actual = actuals_list[month_to - 1] if month_to <= 12 else 0
        total_actual = cumulative_actuals
        progress = (total_actual / annual_target * 100) if annual_target > 0 else 0
        variance = total_actual - annual_target

        table_rows.append({
            'year': year,
            'project': indicator.project.name if indicator.project else "-",
            'indicator': indicator.name,
            'unit': indicator.unit,
            'status': "Active" if indicator.is_active else "Inactive",
            'cumulative': cumulative_actuals,
            'actual': current_actual,
            'total_actual': total_actual,
            'target': annual_target,
            'progress': progress,
            'variance': variance,
        })

    table_rows = sorted(table_rows, key=lambda x: (x['project'], x['indicator']))

    # --- Export Excel ---
    if request.GET.get('export') == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = f"Actual vs Variance {year}"

        headers = ['Year', 'Project', 'Indicator', 'Unit', 'Status', 'Cumulative Actual',
                   'Current Month Actual', 'Total Actual', 'Target (Annual)', 'Variance', 'Progress (%)']
        ws.append(headers)

        header_font = Font(bold=True)
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))

        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
            cell = col[0]
            cell.font = header_font
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border_style
            ws.column_dimensions[cell.column_letter].width = 20

        for row in table_rows:
            variance_val = f"✅ {row['variance']}" if row['variance'] >= 0 else f"❌ {row['variance']}"
            progress_val = f"{round(row['progress'],1)}%"
            row_data = [
                row['year'], row['project'], row['indicator'], row['unit'], row['status'],
                row['cumulative'], row['actual'], row['total_actual'],
                row['target'], variance_val, progress_val
            ]
            ws.append(row_data)

        for r_idx, row_cells in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            for c_idx, cell in enumerate(row_cells, start=1):
                cell.alignment = Alignment(horizontal="left")
                cell.border = border_style
                if c_idx == 10:
                    cell.font = Font(color="008000" if table_rows[r_idx - 2]['variance'] >= 0 else "FF0000")

        footer_row = ws.max_row + 2
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=11)
        footer_cell = ws.cell(row=footer_row, column=1)
        footer_cell.value = f"Actual vs Variance Report generated for year {year}"
        footer_cell.alignment = Alignment(horizontal="center")
        footer_cell.font = Font(italic=True, color="555555")

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=Actual_vs_Variance_{year}.xlsx'
        wb.save(response)
        return response

    # --- Export PDF ---
    if request.GET.get('export') == 'pdf':
        context = {'table_rows': table_rows, 'page_title': f"Actual vs Variance Report {year}"}
        html_string = render(request, 'core/actual_vs_variance_pdf.html', context).content.decode('utf-8')
        pdf_file = BytesIO()
        HTML(string=html_string).write_pdf(pdf_file)
        response = HttpResponse(pdf_file.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=Actual_vs_Variance_{year}.pdf'
        return response

    # --- Chart data ---
    grouped_data = {}
    for indicator in indicators_qs:
        key = (indicator.project.name if indicator.project else "-", indicator.unit)
        target_obj = indicator.targets.filter(year=year).first()
        annual_target = target_obj.value if target_obj else 0
        monthly_target = round(annual_target / 12, 2) if annual_target else 0

        if key not in grouped_data:
            grouped_data[key] = {'project': key[0], 'unit': key[1], 'actuals': [0]*12, 'chart_targets': [0]*12}

        ind_entries = entries.filter(indicator=indicator)
        for entry in ind_entries:
            grouped_data[key]['actuals'][entry.month-1] += entry.value or 0
            grouped_data[key]['chart_targets'][entry.month-1] += monthly_target

    month_labels = [calendar.month_abbr[m] for m in range(month_from, month_to+1)]
    chart_data = []
    for data in grouped_data.values():
        actuals_slice = data['actuals'][month_from-1:month_to]
        targets_slice = data['chart_targets'][month_from-1:month_to]

        chart_data.append({
            'label': f"{data['project']} ({data['unit']}) - Actual",
            'data': actuals_slice,
            'type': 'bar',
            'backgroundColor': 'rgba(54, 162, 235, 0.7)',
        })
        chart_data.append({
            'label': f"{data['project']} ({data['unit']}) - Target",
            'data': targets_slice,
            'type': 'line',
            'borderColor': 'rgba(255, 99, 132, 1)',
            'borderWidth': 2,
            'fill': False,
        })

    context = {
        'year': year,
        'month_from': month_from,
        'month_to': month_to,
        'month_choices': month_choices,
        'indicators': indicators_qs,
        'projects': projects_qs,
        'units': list(entries.values_list('indicator__unit', flat=True).distinct()),
        'selected_indicator': int(indicator_id) if indicator_id else None,
        'selected_unit': unit if unit else None,
        'selected_project': int(project_id) if project_id else None,
        'selected_status': status,
        'categories': json.dumps(month_labels),
        'chart_data': json.dumps(chart_data),
        'table_rows': table_rows,
        'page_title': "📊 Actual vs Variance Report",
    }

    return render(request, 'core/actual_vs_variance.html', context)


from .models import Donor, Project, Facility, Indicator, MonthlyEntry
from django.db.models import Q
# -------------------------
# Donor Detail View
# -------------------------
def donor_detail(request, donor_id):
    donor = get_object_or_404(Donor, id=donor_id)

    # Fetch all projects linked to this donor (main or additional)
    projects = Project.objects.filter(
        Q(main_donor=donor) | Q(donors=donor)
    ).distinct().order_by('start_date')

    # Determine overall start and end dates from linked projects
    if projects.exists():
        start_date = projects.first().start_date
        end_date = projects.last().end_date
    else:
        start_date = None
        end_date = None

    # Initialize donor-level summary counters
    donor_total_kpis = 0
    donor_active_kpis = 0

    # Calculate status and KPI details for each project
    today = date.today()
    projects_with_status = []
    for proj in projects:
        # Determine project status
        if proj.end_date:
            if proj.end_date < today:
                status_text = "Completed"
            elif (proj.end_date - today).days <= 30:
                status_text = "Ending Soon"
            else:
                status_text = "Active"
        else:
            status_text = "Active"

        if proj.end_date and proj.end_date < today and not proj.is_active:
            status_text = "Overdue"

        # Include project facilities for display
        project_facilities = proj.facilities.all()

        # Calculate KPIs
        indicators = Indicator.objects.filter(project=proj)
        total_kpis = indicators.count()

        # Determine active KPIs based on project status and individual KPI status
        if not proj.is_active or status_text in ["Completed", "Overdue"]:
            active_indicators = indicators.none()  # all KPIs inactive
        else:
            active_indicators = indicators.filter(is_active=True)

        active_count = active_indicators.count()
        kpi_url = f"/projects/{proj.id}/kpis/" if total_kpis > 0 else "#"

        # Update donor-level summary
        donor_total_kpis += total_kpis
        donor_active_kpis += active_count

        projects_with_status.append({
            "project": proj,
            "status_text": status_text,
            "facilities": project_facilities,
            "total_kpis": total_kpis,
            "active_count": active_count,
            "kpi_url": kpi_url,
        })

    # Calculate inactive KPIs at donor level
    donor_inactive_kpis = donor_total_kpis - donor_active_kpis

    # Fetch donor's directly linked facilities
    donor_facilities = donor.facilities.all()

    # Include all facilities for map, cast Decimal to float
    all_facilities = Facility.objects.all()
    facilities_json = json.dumps([
        {
            "name": f.name,
            "lat": float(f.latitude),
            "lng": float(f.longitude),
        } for f in all_facilities
    ])

    # -------------------------
    # KPI Trend Data (last 10 months)
    # -------------------------
    trend_labels = []
    trend_total = []
    trend_active = []
    trend_inactive = []

    for i in range(9, -1, -1):
        # Calculate target month and year
        target_month = (today.month - i - 1) % 12 + 1
        target_year = today.year - ((today.month - i - 1) // 12)

        # Get all MonthlyEntry objects for linked projects for this month
        month_entries = MonthlyEntry.objects.filter(
            indicator__project__in=projects,
            year=target_year,
            month=target_month
        )

        month_active = month_entries.filter(indicator__is_active=True)
        month_inactive = month_entries.filter(indicator__is_active=False)

        trend_labels.append(datetime.date(target_year, target_month, 1).strftime("%b"))
        trend_total.append(month_entries.count())
        trend_active.append(month_active.count())
        trend_inactive.append(month_inactive.count())

    # -------------------------
    # Context
    # -------------------------
    context = {
        "donor": donor,
        "projects_with_status": projects_with_status,
        "start_date": start_date,
        "end_date": end_date,
        "today": today,
        "donor_facilities": donor_facilities,
        "facilities_json": facilities_json,
        "donor_total_kpis": donor_total_kpis,
        "donor_active_kpis": donor_active_kpis,
        "donor_inactive_kpis": donor_inactive_kpis,
        "kpi_trend_labels": trend_labels,
        "kpi_trend_total": trend_total,
        "kpi_trend_active": trend_active,
        "kpi_trend_inactive": trend_inactive,
    }

    return render(request, "core/donor_detail.html", context)

# -------------------------
# Donors List View
# -------------------------
from .models import Donor, Project

def donors_list(request):
    """
    Display all donors, including those with no projects.
    Each donor includes their linked projects and project status.
    """
    today = date.today()
    donors = Donor.objects.all().order_by('name')

    donors_data = []

    for donor in donors:
        # Get all projects linked to this donor (may be empty)
        projects = donor.projects.all().order_by('start_date')

        projects_with_status = []
        for proj in projects:
            if proj.end_date:
                if proj.end_date < today:
                    status = "Completed"
                elif (proj.end_date - today).days <= 30:
                    status = "Ending Soon"
                else:
                    status = "Active"
            else:
                status = "Active"

            if proj.end_date and proj.end_date < today and not proj.is_active:
                status = "Overdue"

            projects_with_status.append({
                "project": proj,
                "status": status
            })

        donors_data.append({
            "donor": donor,
            "projects_with_status": projects_with_status  # can be empty
        })

    context = {
        "donors_data": donors_data,
        "today": today,
    }

    return render(request, "core/donors_list.html", context)


from .models import Donor, Project
from .forms import DonorForm

# -------------------------
# Donor Edit View
# -------------------------
def donor_edit(request, donor_id):
    donor = get_object_or_404(Donor, id=donor_id)

    if request.method == "POST":
        form = DonorForm(request.POST, instance=donor)
        if form.is_valid():
            donor = form.save()

            # Update linked projects via the M2M field
            projects_selected = form.cleaned_data.get('projects_from_form')
            if projects_selected is not None:
                donor.projects_from_form.set(projects_selected)

            # Update linked facilities via the new M2M field
            facilities_selected = form.cleaned_data.get('facilities')
            if facilities_selected is not None:
                donor.facilities.set(facilities_selected)

            messages.success(request, f"Donor '{donor.name}' has been updated successfully.")
            return redirect("donor_detail", donor_id=donor.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = DonorForm(instance=donor)

    all_projects = Project.objects.all().order_by('name')
    all_facilities = donor.facilities.all()  # for reference if needed in template

    context = {
        "form": form,
        "donor": donor,
        "all_projects": all_projects,
        "all_facilities": all_facilities,
        "title": "Edit",
    }
    return render(request, "core/donor_form.html", context)


from .models import Donor, Project
from .forms import DonorForm

# -------------------------
# Donor Add View
# -------------------------
def donor_add(request):
    """
    Display a form to create a new donor.
    On POST, validates and saves the donor with optional linked projects and facilities.
    """
    if request.method == "POST":
        form = DonorForm(request.POST)
        if form.is_valid():
            donor = form.save()

            # Link selected projects via the M2M field
            projects_selected = form.cleaned_data.get('projects_from_form')
            if projects_selected is not None:
                donor.projects_from_form.set(projects_selected)

            # Link selected facilities via the new M2M field
            facilities_selected = form.cleaned_data.get('facilities')
            if facilities_selected is not None:
                donor.facilities.set(facilities_selected)

            messages.success(request, f"Donor '{donor.name}' was added successfully!")
            return redirect('donor_detail', donor_id=donor.id)
        else:
            messages.error(request, "Please fix the errors below.")
    else:
        form = DonorForm()

    all_projects = Project.objects.all().order_by('name')
    all_facilities = donor.facilities.all() if 'donor' in locals() and donor else []

    context = {
        "form": form,
        "all_projects": all_projects,
        "all_facilities": all_facilities,
        "donor": None,
        "title": "Add",
    }
    return render(request, "core/donor_form.html", context)


# -------------------------
# Donor Delete View
# -------------------------
def donor_delete(request, donor_id):
    """
    Safely delete a donor after confirmation.
    """
    donor = get_object_or_404(Donor, id=donor_id)

    if request.method == "POST":
        donor_name = donor.name
        donor.delete()
        messages.success(request, f"Donor '{donor_name}' has been deleted successfully.")
        return redirect('donors_list')

    # Render a confirmation page
    return render(request, "core/donor_confirm_delete.html", {"donor": donor})




# Bulk actions in KPIs delete

def project_bulk_action(request, project_id):
    """
    Handles bulk actions on project KPIs: approve, delete, submit.
    Expects POST request with 'action' and 'ids' (list of indicator IDs).
    """
    if request.method != "POST":
        messages.error(request, "Invalid request method.")
        return redirect("project_kpis", pk=project_id)

    project = get_object_or_404(Project, id=project_id)
    action = request.POST.get("action")
    ids = request.POST.getlist("ids")

    if not action or not ids:
        messages.error(request, "No bulk action or rows selected.")
        return redirect("project_kpis", pk=project_id)

    # Fetch only indicators belonging to this project
    indicators = Indicator.objects.filter(id__in=ids, project=project)
    if not indicators.exists():
        messages.error(request, "Selected KPIs do not exist or do not belong to this project.")
        return redirect("project_kpis", pk=project_id)

    try:
        with transaction.atomic():
            if action == "approve":
                updated_count = indicators.update(is_approved=True)
                messages.success(request, f"{updated_count} KPI(s) approved successfully.")
            elif action == "delete":
                deleted_count = indicators.count()
                indicators.delete()
                messages.success(request, f"{deleted_count} KPI(s) deleted successfully.")
            elif action == "submit":
                updated_count = indicators.update(is_submitted=True)
                messages.success(request, f"{updated_count} KPI(s) submitted successfully.")
            else:
                messages.error(request, "Invalid bulk action selected.")
    except Exception as e:
        messages.error(request, f"An error occurred during bulk action: {str(e)}")

    return redirect("project_kpis", pk=project_id)

#logged_out
from django.contrib.auth import logout
from django.shortcuts import redirect
from django.views.decorators.csrf import csrf_exempt

def custom_logout(request):
    if request.method == "POST":
        logout(request)
    return redirect('login')


@login_required
def gender_insights(request):
    import datetime, json, re
    from django.db.models import Sum, Q
    from django.core.cache import cache
    from .utils import safe_int

    today_year = datetime.date.today().year

    # ---------- Request params ----------
    selected_year = safe_int(request.GET.get("year"), default=today_year)
    selected_project_id = safe_int(request.GET.get("project"), default=0)
    selected_unit = request.GET.get("unit") or ""
    selected_district = request.GET.get("district") or ""
    selected_facility = request.GET.get("facility") or ""
    selected_indicator_id = safe_int(request.GET.get("indicator"), default=0)
    month_from = max(1, min(12, safe_int(request.GET.get("month_from"), default=1)))
    month_to = max(1, min(12, safe_int(request.GET.get("month_to"), default=12)))
    if month_from > month_to:
        month_from, month_to = month_to, month_from

    # ---------- Gender keyword matching ----------
    male_terms = ["male", "males", "man", "men", "boy", "boys"]
    female_terms = ["female", "females", "woman", "women", "girl", "girls", "agyw", "agyws"]
    male_regex = r'\b(' + r'|'.join(re.escape(t) for t in male_terms) + r')\b'
    female_regex = r'\b(' + r'|'.join(re.escape(t) for t in female_terms) + r')\b'

    # ---------- Gender-related indicators ----------
    indicators = Indicator.objects.filter(project__is_active=True, is_active=True)
    gender_indicators = indicators.filter(Q(name__iregex=male_regex) | Q(name__iregex=female_regex))

    # ---------- Apply filters ----------
    filtered_indicators = gender_indicators
    if selected_project_id:
        filtered_indicators = filtered_indicators.filter(project__id=selected_project_id)
    if selected_unit:
        filtered_indicators = filtered_indicators.filter(unit__iexact=selected_unit)
    if selected_indicator_id:
        filtered_indicators = filtered_indicators.filter(id=selected_indicator_id)

    all_indicators = filtered_indicators.select_related("project").order_by("project__name", "name")

    # ---------- Filters: full lists ----------
    all_projects = Project.objects.filter(
        is_active=True,
        id__in=gender_indicators.values_list('project_id', flat=True)
    ).distinct()
    all_units = sorted(set(gender_indicators.values_list('unit', flat=True)))
    all_indicators_list = gender_indicators.select_related("project").order_by("name")
    all_districts = sorted(set(District.objects.values_list('name', flat=True)))
    all_facilities = sorted(set(Facility.objects.values_list('name', flat=True)))

    # ---------- Monthly entries ----------
    base_entries = MonthlyEntry.objects.filter(
        year=selected_year,
        indicator__in=all_indicators,
        month__gte=month_from,
        month__lte=month_to
    )

    # Apply filters
    if selected_facility:
        entries = (
            base_entries.filter(indicator__project__facilities__name=selected_facility)
            .values("indicator_id", "year", "month")
            .annotate(value=Sum("value"))
        )
    elif selected_district:
        indicator_ids = (
            Indicator.objects.filter(
                project__facilities__district__name=selected_district,
                project__is_active=True,
                is_active=True,
            )
            .values_list("id", flat=True)
            .distinct()
        )
        entries = (
            MonthlyEntry.objects.filter(
                year=selected_year,
                month__gte=month_from,
                month__lte=month_to,
                indicator_id__in=indicator_ids,
            )
            .values("indicator_id", "year", "month")
            .annotate(value=Sum("value"))
        )
    else:
        entries = base_entries.values("indicator_id", "year", "month").annotate(value=Sum("value"))

    # ---------- Targets ----------
    targets = IndicatorTarget.objects.filter(year=selected_year, indicator__in=all_indicators)
    if selected_facility:
        targets = targets.filter(indicator__project__facilities__name=selected_facility)
    elif selected_district:
        targets = targets.filter(indicator__project__facilities__district__name=selected_district)

    # ---------- Pivot & indicator totals ----------
    pivot_rows = []
    months_range = list(range(month_from, month_to + 1))
    month_labels = [datetime.date(1900, m, 1).strftime("%b") for m in months_range]

    total_male = total_female = 0
    per_indicator_data = []

    entries_dict = {}
    for e in entries:
        key = (e["indicator_id"], e["month"])
        entries_dict[key] = entries_dict.get(key, 0) + e["value"]

    for ind in all_indicators:
        row = {
            "year": selected_year,
            "project": ind.project.name if ind.project else "—",
            "indicator": ind.name,
            "unit": ind.unit or "",
            "months": [],
            "actual_total": 0,
            "target_total": 0,
            "percent": 0,
        }
        male_val = female_val = 0
        for m in months_range:
            month_total = entries_dict.get((ind.id, m), 0)
            row["months"].append(month_total)
            row["actual_total"] += month_total
            if re.search(male_regex, ind.name, re.IGNORECASE):
                male_val += month_total
            elif re.search(female_regex, ind.name, re.IGNORECASE):
                female_val += month_total
        target_total = sum(t.value for t in targets.filter(indicator=ind))
        row["target_total"] = target_total
        row["percent"] = round((row["actual_total"] / target_total * 100) if target_total else 0, 1)
        pivot_rows.append(row)

        per_indicator_data.append({
            "name": ind.name,
            "male_value": male_val,
            "female_value": female_val,
            "districts": list(ind.project.facilities.values_list("district__name", flat=True))
        })
        total_male += male_val
        total_female += female_val

    total_actual = total_male + total_female

    # ---------- Summary metrics ----------
    gender_summary = {
        "male_value": total_male,
        "female_value": total_female,
        "male_percent": round((total_male / total_actual * 100), 1) if total_actual else 0,
        "female_percent": round((total_female / total_actual * 100), 1) if total_actual else 0,
    }
    gender_gap = abs(gender_summary["male_percent"] - gender_summary["female_percent"])
    dominant_gender = (
        "Male" if total_male > total_female else
        "Female" if total_female > total_male else
        "Equal"
    )

    # ---------- Monthly chart data ----------
    male_monthly, female_monthly = [], []
    for m in months_range:
        male_val = sum(
            entries_dict.get((ind.id, m), 0)
            for ind in all_indicators if re.search(male_regex, ind.name, re.IGNORECASE)
        )
        female_val = sum(
            entries_dict.get((ind.id, m), 0)
            for ind in all_indicators if re.search(female_regex, ind.name, re.IGNORECASE)
        )
        male_monthly.append(male_val)
        female_monthly.append(female_val)

    # ---------- Facility Map Data ----------
    facility_data = []
    active_facility_names = []
    facilities_qs = Facility.objects.exclude(latitude__isnull=True).exclude(longitude__isnull=True)
    if selected_facility:
        facilities_qs = facilities_qs.filter(name=selected_facility)
    elif selected_district:
        facilities_qs = facilities_qs.filter(district__name=selected_district)

    for f in facilities_qs:
        male_val = sum(d["male_value"] for d in per_indicator_data if f.district.name in d["districts"])
        female_val = sum(d["female_value"] for d in per_indicator_data if f.district.name in d["districts"])
        total = male_val + female_val
        if total == 0:
            continue
        female_percent = round((female_val / total * 100), 1)
        facility_data.append({
            "facility": f.name,
            "district": f.district.name if f.district else "—",
            "male_value": male_val,
            "female_value": female_val,
            "total": total,
            "female_percent": female_percent,
            "lat": float(f.latitude),
            "lng": float(f.longitude),
            "icon": "/static/images/marker-icon.png",
        })
        active_facility_names.append(f.name)

    # ---------- District map totals using pivot data ----------
    district_map_data = {}
    district_totals = {}
    for d in per_indicator_data:
        for dist in d["districts"]:
            if dist not in district_map_data:
                district_map_data[dist] = {"male": 0, "female": 0}
            district_map_data[dist]["male"] += d["male_value"]
            district_map_data[dist]["female"] += d["female_value"]

    final_district_map_data = []
    for name, data in district_map_data.items():
        male_val = data["male"]
        female_val = data["female"]
        total = male_val + female_val
        if total == 0:
            continue
        female_percent = round((female_val / total * 100), 1)
        final_district_map_data.append({
            "district": name,
            "female": female_val,
            "male": male_val,
            "total": total,
            "female_percent": female_percent,
        })
        district_totals[name] = female_val

    # ---------- Bubble Chart Data ----------
    bubble_data = []
    for d in per_indicator_data:
        x_val, y_val = d["male_value"], d["female_value"]
        if x_val == 0 and y_val == 0:
            continue
        r_val = max(5, (x_val + y_val) / 10)
        bubble_data.append({
            "indicator": d["name"],
            "male": x_val,
            "female": y_val,
            "size": r_val
        })

    # ---------- Chart Data JSON ----------
    chart_data_json = json.dumps({
        "month_labels": month_labels,
        "male_monthly": male_monthly,
        "female_monthly": female_monthly,
        "gender_summary": gender_summary,
        "facility_data": facility_data,
        "bubble_data": bubble_data,
        "district_map_data": final_district_map_data,
        "active_facility_names": active_facility_names,
    })

    # ---------- Filters ----------
    years_set = set(
        list(MonthlyEntry.objects.values_list("year", flat=True).distinct()) +
        list(IndicatorTarget.objects.values_list("year", flat=True).distinct()) +
        [today_year]
    )
    years = sorted(years_set, reverse=True)
    months_full = list(range(1, 13))

    # ---------- Determine top district ----------
    top_district = max(district_totals.items(), key=lambda x: x[1])[0] if district_totals else "All Districts"

    context = {
        "title": "Gender Insights",
        "years": years,
        "selected_year": selected_year,
        "projects": all_projects,
        "selected_project_id": selected_project_id,
        "units": all_units,
        "selected_unit": selected_unit,
        "districts": all_districts,
        "selected_district": selected_district,
        "facilities": all_facilities,
        "selected_facility": selected_facility,
        "indicators": all_indicators_list,
        "selected_indicator_id": selected_indicator_id,
        "months": months_full,
        "month_from": month_from,
        "month_to": month_to,
        "month_labels": month_labels,
        "pivot_rows": pivot_rows,
        "chart_data_json": chart_data_json,
        "gender_summary": gender_summary,
        "gender_gap": gender_gap,
        "dominant_gender": dominant_gender,
        "top_district": top_district,
        "geojson_path": "/static/geojson/geojsonmalawi_districts.geojson",
    }

    return render(request, "core/gender_insights.html", context)



@login_required
def download_gender_excel(request):
    """
    Download Excel pivot table for Gender Insights
    """
    import io
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.db.models import Q, Sum
    import datetime, re

    today_year = datetime.date.today().year
    selected_year = int(request.GET.get("year", today_year))
    selected_project_id = int(request.GET.get("project", 0))
    selected_unit = request.GET.get("unit", "")
    selected_indicator_id = int(request.GET.get("indicator", 0))

    # Gender regex
    male_terms = ["male", "males", "man", "men", "boy", "boys"]
    female_terms = ["female", "females", "woman", "women", "girl", "girls", "agyw", "agyws"]
    male_regex = r'\b(' + r'|'.join(re.escape(t) for t in male_terms) + r')\b'
    female_regex = r'\b(' + r'|'.join(re.escape(t) for t in female_terms) + r')\b'

    # Filter indicators
    indicators = Indicator.objects.filter(project__is_active=True, is_active=True)
    gender_indicators = indicators.filter(Q(name__iregex=male_regex) | Q(name__iregex=female_regex))
    if selected_project_id:
        gender_indicators = gender_indicators.filter(project__id=selected_project_id)
    if selected_unit:
        gender_indicators = gender_indicators.filter(unit__iexact=selected_unit)
    if selected_indicator_id:
        gender_indicators = gender_indicators.filter(id=selected_indicator_id)
    all_indicators = gender_indicators.select_related("project").order_by("project__name", "name")

    months = list(range(1, 13))
    month_labels = [datetime.date(1900, m, 1).strftime("%b") for m in months]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Gender Pivot Table"

    # Styles
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left")
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    # --- Title ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(months) + 3)
    ws.cell(row=1, column=1, value="Life Concern Data Management System - Gender Insights").font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = center

    # --- Filters info ---
    ws.cell(row=2, column=1, value=f"Year: {selected_year}").alignment = left
    ws.cell(row=2, column=2, value=f"Project: {selected_project_id or 'All'}").alignment = left
    ws.cell(row=2, column=3, value=f"Unit: {selected_unit or 'All'}").alignment = left
    ws.cell(row=2, column=4, value=f"Indicator: {selected_indicator_id or 'All'}").alignment = left

    # --- Headers ---
    headers = ["Year", "Project", "Indicator", "Unit"] + month_labels + ["Actual Total", "Target Total", "% Achieved"]
    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = bold
        cell.alignment = center
        cell.border = border
        cell.fill = header_fill

    # --- Data rows ---
    row_idx = header_row + 1
    for ind in all_indicators:
        ws.cell(row=row_idx, column=1, value=selected_year).alignment = left
        ws.cell(row=row_idx, column=2, value=ind.project.name if ind.project else "—").alignment = left
        ws.cell(row=row_idx, column=3, value=ind.name).alignment = left
        ws.cell(row=row_idx, column=4, value=ind.unit or "").alignment = left

        actual_total = 0
        col = 5
        for m in months:
            val = MonthlyEntry.objects.filter(year=selected_year, indicator=ind, month=m).aggregate(total=Sum("value"))["total"] or 0
            ws.cell(row=row_idx, column=col, value=int(val)).alignment = center
            ws.cell(row=row_idx, column=col).border = border
            actual_total += int(val)
            col += 1

        target_total = IndicatorTarget.objects.filter(year=selected_year, indicator=ind).aggregate(total=Sum("value"))["total"] or 0
        percent = round((actual_total / target_total * 100) if target_total else 0, 1)

        ws.cell(row=row_idx, column=col, value=actual_total).alignment = center
        ws.cell(row=row_idx, column=col).border = border
        ws.cell(row=row_idx, column=col+1, value=target_total).alignment = center
        ws.cell(row=row_idx, column=col+1).border = border
        ws.cell(row=row_idx, column=col+2, value=percent).alignment = center
        ws.cell(row=row_idx, column=col+2).border = border

        row_idx += 1

    # --- Footer ---
    footer_row = row_idx + 1
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=4 + len(months) + 3)
    ws.cell(row=footer_row, column=1, value="Life Concern Data Management System").alignment = center
    ws.cell(footer_row, 1).font = Font(italic=True, color="888888")

    # --- Auto column width ---
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    # --- Save workbook to response ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="Gender_Pivot_Table.xlsx"'
    return response


